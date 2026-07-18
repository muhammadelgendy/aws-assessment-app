"""
AWS Full Account Assessment Script
Collects data across all services, generates Excel report with:
  - Account Overview sheet
  - Per-service detail sheets
  - Recommendations & Gaps sheet (colour-coded by severity)

Run with: python lambda_function_fi.py
"""

import boto3
import pandas as pd
from datetime import datetime, timedelta, timezone
import json
import io
import os
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.pdfgen.canvas import Canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


# ── Config ─────────────────────────────────────────────────────────────────────
TARGET_REGION  = None   # None = all regions, e.g. 'eu-west-1' for one region
EXCLUDED_REGIONS = [
    'me-south-1',           # Middle East
    # 'ap-northeast-1',       # Asia Pacific - Tokyo
    # 'ap-northeast-2',       # Asia Pacific - Seoul
    # 'ap-northeast-3',       # Asia Pacific - Osaka
    # 'ap-southeast-1',       # Asia Pacific - Singapore
    # 'ap-southeast-2',       # Asia Pacific - Sydney
    # 'ap-east-1',   
    # 'sa-east-1', 
    # 'eu-west-2',
    # 'ca-central-1',
    # 'eu-west-3' ,   # Asia Pacific - Hong Kong
]
MAX_WORKERS = 40           # global thread-pool ceiling (reduce if hitting throttling)
REGION_WORKERS = 3     # max parallel regions per collector
BUCKET_WORKERS = 50     # max parallel S3 bucket checks
# ──────────────────────────────────────────────────────────────────────────────

COLOUR = {
    'CRITICAL': 'FFB3B3',
    'HIGH':     'FFD9B3',
    'MEDIUM':   'FFFBB3',
    'LOW':      'D9F0B3',
    'INFO':     'D6EAF8',
}

def _now() -> datetime:
    return datetime.now(timezone.utc)


class AWSAssessment:
    def __init__(self, profile_name=None, aws_access_key_id=None, aws_secret_access_key=None,
                 aws_session_token=None, region_name=None):
        if aws_access_key_id and aws_secret_access_key:
            # Credentials submitted directly via the web form
            self.session = boto3.Session(
                aws_access_key_id=aws_access_key_id,
                aws_secret_access_key=aws_secret_access_key,
                aws_session_token=aws_session_token,
                region_name=region_name,
            )
        elif profile_name:
            self.session = boto3.Session(profile_name=profile_name)
        else:
            self.session = boto3.Session()
        self.regions      = self._get_regions()
        self.report_data  = {}
        self.findings     = []
        self._lock        = threading.Lock()
        self._clients     = {}          # (service, region) → client cache
        self._client_lock = threading.Lock()
        self.account_id = self._client('sts', 'us-east-1').get_caller_identity()['Account']
        self._account_name = self._get_account_name()   # call it once here

        try:
            self.account_id = self._client('sts', 'us-east-1').get_caller_identity()['Account']
        except Exception:
            self.account_id = 'Unknown'
        self._account_name = self._get_account_name()
    @property
    def account_name(self):
        """Return cached account name (set in __init__)."""
        return self._account_name

    # ── Client cache (avoids rebuilding the same client repeatedly) ────────────
    def _client(self, service, region='us-east-1'):
        key = (service, region)
        with self._client_lock:
            if key not in self._clients:
                self._clients[key] = self.session.client(service, region_name=region)
            return self._clients[key]

    # ── Run a per-region function across ALL regions in parallel ───────────────
    def _run_in_regions(self, fn, workers=None):
        """
        fn(region) → list[dict]   (must be thread-safe, uses self._client())
        Returns flat list of all results merged from every region.
        """
        results = []
        w = workers or min(REGION_WORKERS, len(self.regions))
        with ThreadPoolExecutor(max_workers=w) as ex:
            futures = {ex.submit(fn, r): r for r in self.regions}
            for future in as_completed(futures):
                try:
                    chunk = future.result()
                    if chunk:
                        results.extend(chunk)
                except Exception as e:
                    print(f"    ⚠ region error: {e}")
        return results

    # ── Helpers ────────────────────────────────────────────────────────────────
    def _get_regions(self):
        if TARGET_REGION:
            print(f"Region scope: {TARGET_REGION}")
            return [TARGET_REGION]
        print("Region scope: ALL regions (except excluded)")
        ec2 = self.session.client('ec2', region_name='us-east-1')
        all_regions = [r['RegionName'] for r in ec2.describe_regions()['Regions']]
        # Filter out excluded regions
        filtered_regions = [r for r in all_regions if r not in EXCLUDED_REGIONS]
        if EXCLUDED_REGIONS:
            print(f"  Excluded regions: {', '.join(EXCLUDED_REGIONS)}")
        return filtered_regions

    def _naive(self, dt):
        if dt is None:
            return None
        if hasattr(dt, 'tzinfo') and dt.tzinfo is not None:
            return dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt

    def _get_account_name(self):
        try:
            iam = self._client('iam', 'us-east-1')
            aliases = iam.list_account_aliases()['AccountAliases']
            return aliases[0] if aliases else 'N/A'
        except Exception:
            return 'N/A'

    def _add_finding(self, severity, category, resource, issue, recommendation):
        with self._lock:
            self.findings.append({
                'Severity':       severity,
                'Category':       category,
                'Resource':       resource,
                'Issue':          issue,
                'Recommendation': recommendation,
            })

    def _store(self, key, data):
        with self._lock:
            self.report_data[key] = data

    # ─────────────────────────────────────────────────────────────────────────
    # COLLECTORS  (each one fans out across regions in parallel internally)
    # ─────────────────────────────────────────────────────────────────────────

    # ── 1. IAM Users (parallel per-user API calls) ────────────────────────────
    def get_iam_users_data(self):
        print("Collecting IAM users...")
        iam  = self._client('iam', 'us-east-1')
        now  = _now()
        users = [u for page in iam.get_paginator('list_users').paginate() for u in page['Users']]

        def _enrich(user):
            uname = user['UserName']
            mfa   = iam.list_mfa_devices(UserName=uname)['MFADevices']
            keys  = iam.list_access_keys(UserName=uname)['AccessKeyMetadata']
            pwd   = user.get('PasswordLastUsed')
            days  = (now - pwd).days if pwd else None
            status = 'Inactive' if days and days > 90 else ('Never Active' if pwd is None else 'Active')
            old_k  = [k for k in keys if (now - k['CreateDate']).days > 90]

            if not mfa:
                self._add_finding('HIGH', 'IAM', uname, 'MFA not enabled',
                    'Enable MFA for all IAM users with console access.')
            if status == 'Inactive':
                self._add_finding('MEDIUM', 'IAM', uname, f'Inactive for {days} days',
                    'Remove or disable users inactive for >90 days.')
            if old_k:
                self._add_finding('MEDIUM', 'IAM', uname, f'{len(old_k)} access key(s) >90 days old',
                    'Rotate access keys every 90 days; prefer IAM roles.')

            return {
                'User Name':          uname,
                'ARN':                user['Arn'],
                'Create Date':        self._naive(user['CreateDate']),
                'MFA Enabled':        'Yes' if mfa else 'No',
                'Status':             status,
                'Password Last Used': self._naive(pwd) if pwd else 'Never',
                'Days Since Login':   days if days is not None else 'Never',
                'Access Keys':        len(keys),
                'Old Keys (>90d)':    len(old_k),
            }

        data = []
        # Each user needs 2 extra API calls — parallelise them
        with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(users) or 1)) as ex:
            for result in as_completed([ex.submit(_enrich, u) for u in users]):
                try:
                    data.append(result.result())
                except Exception as e:
                    print(f"    ⚠ IAM user error: {e}")

        self._store('IAM Users', data)
        return len(data)

    # ── 2. IAM Password Policy (global, instant) ───────────────────────────────
    def get_iam_password_policy(self):
        print("Collecting IAM password policy...")
        iam  = self._client('iam', 'us-east-1')
        data = []
        try:
            p = iam.get_account_password_policy()['PasswordPolicy']
            data.append({
                'Min Length':             p.get('MinimumPasswordLength', 'N/A'),
                'Require Uppercase':      p.get('RequireUppercaseCharacters', False),
                'Require Lowercase':      p.get('RequireLowercaseCharacters', False),
                'Require Numbers':        p.get('RequireNumbers', False),
                'Require Symbols':        p.get('RequireSymbols', False),
                'Max Age (days)':         p.get('MaxPasswordAge', 'No expiry'),
                'Password Reuse Prevent': p.get('PasswordReusePrevention', 'None'),
                'Allow Users to Change':  p.get('AllowUsersToChangePassword', False),
                'Hard Expiry':            p.get('HardExpiry', False),
            })
            if p.get('MinimumPasswordLength', 0) < 14:
                self._add_finding('MEDIUM', 'IAM', 'Password Policy',
                    f"Min length {p.get('MinimumPasswordLength')} (recommended ≥14)",
                    'Set minimum password length to at least 14.')
            if not p.get('RequireSymbols'):
                self._add_finding('MEDIUM', 'IAM', 'Password Policy', 'Symbols not required',
                    'Enable symbol requirement in account password policy.')
            if not p.get('MaxPasswordAge') or p.get('MaxPasswordAge', 999) > 90:
                self._add_finding('LOW', 'IAM', 'Password Policy', 'Password expiry >90 days or not set',
                    'Set password expiry to ≤90 days.')
            if not p.get('PasswordReusePrevention') or p.get('PasswordReusePrevention', 0) < 5:
                self._add_finding('LOW', 'IAM', 'Password Policy', 'Password reuse prevention < 5',
                    'Prevent reuse of at least the last 5 passwords.')
        except iam.exceptions.NoSuchEntityException:
            data.append({'Note': 'No account password policy set'})
            self._add_finding('HIGH', 'IAM', 'Password Policy', 'No password policy configured',
                'Create a strong account password policy.')
        self._store('IAM Password Policy', data)
        return len(data)

    # ── 3. IAM Roles (parallel pagination) ────────────────────────────────────
    def get_iam_roles(self):
        print("Collecting IAM roles...")
        iam   = self._client('iam', 'us-east-1')
        roles = [r for page in iam.get_paginator('list_roles').paginate() for r in page['Roles']]
        data  = []
        for role in roles:
            trust     = json.dumps(role.get('AssumeRolePolicyDocument', {}))
            is_public = '"AWS": "*"' in trust or '"Service": "*"' in trust
            data.append({
                'Role Name':               role['RoleName'],
                'ARN':                     role['Arn'],
                'Create Date':             self._naive(role['CreateDate']),
                'Overly Permissive Trust': 'Yes' if is_public else 'No',
                'Trust Policy (summary)':  trust[:200],
            })
            if is_public:
                self._add_finding('CRITICAL', 'IAM', role['RoleName'],
                    'Trust policy allows any principal ("*")',
                    'Restrict trust policy to specific accounts/services.')
        self._store('IAM Roles', data)
        return len(data)

    # ── 4. EC2 Instances (parallel across regions) ────────────────────────────
    def get_ec2_instances(self):
        print("Collecting EC2 instances...")
        def _region(region):
            rows = []
            ec2  = self._client('ec2', region)
            for res in ec2.describe_instances()['Reservations']:
                for i in res['Instances']:
                    name    = next((t['Value'] for t in i.get('Tags', []) if t['Key'] == 'Name'), 'N/A')
                    imdsv1  = i.get('MetadataOptions', {}).get('HttpTokens', 'optional') == 'optional'
                    rows.append({
                        'Region':        region,
                        'Instance ID':   i['InstanceId'],
                        'Name':          name,
                        'State':         i['State']['Name'],
                        'Instance Type': i['InstanceType'],
                        'Launch Time':   self._naive(i['LaunchTime']),
                        'Public IP':     i.get('PublicIpAddress', 'N/A'),
                        'Private IP':    i.get('PrivateIpAddress', 'N/A'),
                        'VPC ID':        i.get('VpcId', 'N/A'),
                        'IMDSv2 Only':   'No' if imdsv1 else 'Yes',
                        'Monitoring':    i.get('Monitoring', {}).get('State', 'N/A'),
                    })
                    if imdsv1:
                        self._add_finding('MEDIUM', 'EC2', f"{name} ({i['InstanceId']})",
                            'IMDSv1 enabled (SSRF risk)', 'Enforce IMDSv2 (HttpTokens=required).')
                    if i.get('PublicIpAddress') and i['State']['Name'] == 'running':
                        self._add_finding('INFO', 'EC2', f"{name} ({i['InstanceId']})",
                            'Instance has a public IP',
                            'Use a load balancer or NAT GW instead of direct public IPs where possible.')
            return rows

        data = self._run_in_regions(_region)
        data.sort(key=lambda x: x['Launch Time'] or datetime.min, reverse=True)
        self._store('EC2 Instances', data)
        return len(data)
    # ── 5. AMIs (parallel across regions) ──────────────────────────
    def get_amis(self):
        print("Collecting AMIs owned by account (all regions)...")
        now = _now()
        data = []

        def _region(region):
            rows = []
            try:
                ec2 = self._client('ec2', region)
                images = ec2.describe_images(Owners=['self'])['Images']
                for img in images:
                    # Parse CreationDate (string) to datetime
                    creation_str = img.get('CreationDate')
                    if creation_str:
                        try:
                            # AWS format: 2024-01-15T12:34:56.000Z
                            creation_date = datetime.strptime(creation_str, '%Y-%m-%dT%H:%M:%S.%fZ')
                            creation_date = creation_date.replace(tzinfo=timezone.utc)
                            age_days = (now - creation_date).days
                        except Exception:
                            creation_date = None
                            age_days = 'N/A'
                    else:
                        creation_date = None
                        age_days = 'N/A'

                    name = next((t['Value'] for t in img.get('Tags', []) if t['Key'] == 'Name'), img.get('Name', 'N/A'))
                    rows.append({
                        'Region': region,
                        'AMI ID': img['ImageId'],
                        'Name': name,
                        'Description': img.get('Description', ''),
                        'Creation Date': self._naive(creation_date) if creation_date else 'N/A',
                        'Age (days)': age_days if age_days != 'N/A' else 'N/A',
                        'Public': 'Yes' if img.get('Public') else 'No',
                        'State': img.get('State'),
                        'Platform': img.get('Platform', 'N/A'),
                        'Architecture': img.get('Architecture', 'N/A'),
                        'Tags': str(img.get('Tags', [])),
                    })
                    if isinstance(age_days, int) and age_days > 90:
                        self._add_finding('LOW', 'AMIs', f"{img['ImageId']} ({region})",
                                        f'AMI older than 90 days ({age_days} days)',
                                        'Consider deregistering unused old AMIs.')
            except Exception as e:
                print(f"    ⚠ Error collecting AMIs in {region}: {e}")
            return rows

        data = self._run_in_regions(_region)
        self._store('AMIs', data)
        return len(data)

    # ── 5. Security Groups (parallel across regions) ──────────────────────────
    def get_security_groups(self):
        print("Collecting security groups...")
        def _region(region):
            rows = []
            for sg in self._client('ec2', region).describe_security_groups()['SecurityGroups']:
                open_rules = [
                    r for r in sg.get('IpPermissions', [])
                    for ip in r.get('IpRanges', []) if ip.get('CidrIp') == '0.0.0.0/0'
                ]
                open_ports = [f"{r.get('IpProtocol')}:{r.get('FromPort','*')}-{r.get('ToPort','*')}"
                              for r in open_rules]
                rows.append({
                    'Region':            region,
                    'SG ID':             sg['GroupId'],
                    'Name':              sg['GroupName'],
                    'VPC ID':            sg.get('VpcId', 'N/A'),
                    'Open to 0.0.0.0/0': 'Yes' if open_rules else 'No',
                    'Open Ports':        ', '.join(open_ports) or 'None',
                    'Description':       sg.get('Description', ''),
                })
                for r in open_rules:
                    fp    = r.get('FromPort')
                    proto = r.get('IpProtocol', '-1')
                    if fp in (22, 3389):
                        sev, issue = 'CRITICAL', f'Port {fp} (SSH/RDP) open to 0.0.0.0/0'
                    elif proto == '-1':
                        sev, issue = 'CRITICAL', 'All traffic allowed from 0.0.0.0/0'
                    else:
                        sev, issue = 'HIGH', f'Port {fp} open to 0.0.0.0/0'
                    self._add_finding(sev, 'Security Groups', sg['GroupId'], issue,
                        'Restrict ingress to known IP ranges. Never expose SSH/RDP to the world.')
            return rows

        data = self._run_in_regions(_region)
        self._store('Security Groups', data)
        return len(data)

    # ── 6. EBS Volumes (parallel across regions) ──────────────────────────────
    def get_ebs_volumes(self):
        print("Collecting EBS volumes...")
        def _region(region):
            rows = []
            for v in self._client('ec2', region).describe_volumes()['Volumes']:
                att = v['Attachments']
                enc = v.get('Encrypted', False)
                att_to = att[0]['InstanceId'] if att else 'Unattached'
                # Get volume name from tags
                vol_name = next((t['Value'] for t in v.get('Tags', []) if t['Key'] == 'Name'), 'N/A')
                rows.append({
                    'Region':      region,
                    'Volume ID':   v['VolumeId'],
                    'Volume Name': vol_name,        # <-- new column
                    'Size (GB)':   v['Size'],
                    'Volume Type': v['VolumeType'],
                    'State':       v['State'],
                    'Attached To': att_to,
                    'Encrypted':   'Yes' if enc else 'No',
                    'Create Time': self._naive(v['CreateTime']),
                    'Snapshot ID': v.get('SnapshotId', 'N/A'),
                })
                if not enc:
                    self._add_finding('HIGH', 'EBS', v['VolumeId'], 'Volume not encrypted',
                        'Enable EBS encryption by default; re-create unencrypted volumes.')
                if att_to == 'Unattached':
                    self._add_finding('LOW', 'EBS', v['VolumeId'], 'Unattached volume (idle cost)',
                        'Delete or snapshot unattached EBS volumes.')
            return rows

        data = self._run_in_regions(_region)
        self._store('EBS Volumes', data)
        return len(data)

    # ── 7. EBS Snapshots (parallel across regions) ────────────────────────────
    def get_ebs_snapshots(self):
        print("Collecting EBS snapshots...")
        now = _now()
        ninety_ago = now - timedelta(days=90)

        def _region(region):
            rows = []
            for s in self._client('ec2', region).describe_snapshots(OwnerIds=['self'])['Snapshots']:
                st  = s['StartTime']
                age = (now - st).days
                rows.append({
                    'Region':      region,
                    'Snapshot ID': s['SnapshotId'],
                    'Volume ID':   s.get('VolumeId', 'N/A'),
                    'Size (GB)':   s['VolumeSize'],
                    'Start Time':  self._naive(st),
                    'Encrypted':   'Yes' if s.get('Encrypted') else 'No',
                    'Age (days)':  age,
                    'Status':      'Old (>90d)' if st < ninety_ago else 'Recent',
                    'Description': s.get('Description', ''),
                })
                if not s.get('Encrypted'):
                    self._add_finding('MEDIUM', 'EBS Snapshots', s['SnapshotId'],
                        'Snapshot not encrypted',
                        'Copy snapshot with encryption and delete the unencrypted copy.')
            return rows

        data = self._run_in_regions(_region)
        self._store('EBS Snapshots', data)
        return len(data)

    # ── 8. RDS Instances (parallel regions + parallel CW per instance) ───────────

    def get_rds_instances(self):
        print("Collecting RDS instances (optimized with GetMetricData)...")

        def _region(region):
            rows = []
            rds = self._client('rds', region)
            cw = self._client('cloudwatch', region)

            instances = rds.describe_db_instances()['DBInstances']
            if not instances:
                return rows

            # ---- Batch CPU metrics using GetMetricData ----
            # Build metric queries for each instance
            metric_queries = []
            for idx, inst in enumerate(instances):
                inst_id = inst['DBInstanceIdentifier']
                metric_queries.append({
                    'Id': f'm{idx}',
                    'MetricStat': {
                        'Metric': {
                            'Namespace': 'AWS/RDS',
                            'MetricName': 'CPUUtilization',
                            'Dimensions': [
                                {'Name': 'DBInstanceIdentifier', 'Value': inst_id}
                            ]
                        },
                        'Period': 86400,      # last 24 hours
                        'Stat': 'Average',
                        'Unit': 'Percent'
                    },
                    'ReturnData': True
                })

            # Batch request – one API call for the whole region
            cpu_data = {}
            try:
                # Split into chunks of 500 (safety, though 500 is the hard limit)
                for i in range(0, len(metric_queries), 500):
                    chunk = metric_queries[i:i+500]
                    response = cw.get_metric_data(
                        MetricDataQueries=chunk,
                        StartTime=_now() - timedelta(hours=24),
                        EndTime=_now(),
                        ScanBy='TimestampDescending'
                    )
                    for res in response['MetricDataResults']:
                        # Map back to instance id
                        idx = int(res['Id'][1:])     # 'm0' -> 0
                        inst_id = instances[idx]['DBInstanceIdentifier']
                        avg_cpu = res['Values'][0] if res['Values'] else None
                        cpu_data[inst_id] = round(avg_cpu, 2) if avg_cpu is not None else 'N/A'
            except Exception as e:
                print(f"    ⚠ CloudWatch GetMetricData error in {region}: {e}")
                # fallback: mark all as 'N/A'
                for inst in instances:
                    cpu_data[inst['DBInstanceIdentifier']] = 'N/A'

            # ---- Enrich each instance (now without per-instance CW calls) ----
            for inst in instances:
                inst_id = inst['DBInstanceIdentifier']
                pub = inst.get('PubliclyAccessible', False)
                maz = inst.get('MultiAZ', False)
                enc = inst.get('StorageEncrypted', False)
                bkp = inst.get('BackupRetentionPeriod', 0)
                avg_cpu = cpu_data.get(inst_id, 'N/A')

                # Add findings (same as before)
                if pub:
                    self._add_finding('CRITICAL', 'RDS', inst_id,
                        'RDS publicly accessible', 'Disable public access; place in private subnet.')
                if not enc:
                    self._add_finding('HIGH', 'RDS', inst_id,
                        'RDS storage not encrypted', 'Migrate to an encrypted instance.')
                if not maz:
                    self._add_finding('MEDIUM', 'RDS', inst_id,
                        'Multi-AZ not enabled', 'Enable Multi-AZ for production databases.')
                if bkp < 7:
                    self._add_finding('MEDIUM', 'RDS', inst_id,
                        f'Backup retention only {bkp} day(s)', 'Set backup retention ≥7 days.')

                rows.append({
                    'Region':              region,
                    'DB Identifier':       inst_id,
                    'Engine':              inst['Engine'],
                    'Engine Version':      inst['EngineVersion'],
                    'Status':              inst['DBInstanceStatus'],
                    'Instance Class':      inst['DBInstanceClass'],
                    'Storage (GB)':        inst['AllocatedStorage'],
                    'Multi-AZ':            'Yes' if maz else 'No',
                    'Encrypted':           'Yes' if enc else 'No',
                    'Publicly Accessible': 'Yes' if pub else 'No',
                    'Backup Retention':    bkp,
                    'CPU Util (%)':        avg_cpu,
                    'Endpoint':            inst['Endpoint']['Address'],
                    'Create Time':         self._naive(inst.get('InstanceCreateTime')),
                })
            return rows

        data = self._run_in_regions(_region)
        self._store('RDS Instances', data)
        return len(data)

    # ── 9. S3 Buckets (parallel buckets + parallel calls per bucket) ─────────────
    def get_s3_buckets(self):
        print("Collecting S3 buckets (with region skip & timeout)...")
        from botocore.config import Config
        config = Config(connect_timeout=5, read_timeout=10, retries={'max_attempts': 1})
        s3 = self.session.client('s3', region_name='us-east-1', config=config)
        
        try:
            buckets = s3.list_buckets()['Buckets']
        except Exception as e:
            print(f"  ✗ Error listing buckets: {e}")
            self._store('S3 Buckets', [])
            return 0

        print(f"  → Found {len(buckets)} bucket(s)")
        data = []

        for idx, bucket in enumerate(buckets, 1):
            name = bucket['Name']
            print(f"    [{idx}/{len(buckets)}] Checking '{name}'...", end=' ', flush=True)
            
            # ---- Step 1: Get bucket region (with timeout) ----
            try:
                # Use a short timeout for location to avoid hanging on excluded regions
                with ThreadPoolExecutor(max_workers=1) as ex:
                    loc_future = ex.submit(lambda: s3.get_bucket_location(Bucket=name))
                    loc_response = loc_future.result(timeout=10)
                bucket_region = loc_response.get('LocationConstraint') or 'us-east-1'
                
                # Skip bucket if its region is in EXCLUDED_REGIONS (optional)
                if bucket_region in EXCLUDED_REGIONS:
                    print(f"⏭ Skipped (region {bucket_region} excluded)")
                    continue
            except Exception as e:
                print(f"⚠ Cannot determine region: {e} – skipping")
                continue
            
            # ---- Now safely check the bucket (only if region is not excluded) ----
            try:
                # Re‑create client with region‑specific endpoint (optional but faster)
                s3_bucket = self.session.client('s3', region_name=bucket_region, config=config)
                
                # Encryption
                try:
                    s3_bucket.get_bucket_encryption(Bucket=name)
                    encrypted = 'Yes'
                except Exception:
                    encrypted = 'No'
                
                # Public access block & ACL
                try:
                    pab = s3_bucket.get_public_access_block(Bucket=name)['PublicAccessBlockConfiguration']
                    public_access = 'No' if all(pab.values()) else 'Partial block'
                except Exception:
                    public_access = 'No'
                try:
                    acl = s3_bucket.get_bucket_acl(Bucket=name)
                    for grant in acl['Grants']:
                        uri = grant.get('Grantee', {}).get('URI', '')
                        if 'AllUsers' in uri or 'AuthenticatedUsers' in uri:
                            public_access = 'Yes'
                            break
                except Exception:
                    pass
                
                # Versioning
                try:
                    vs = s3_bucket.get_bucket_versioning(Bucket=name)
                    versioning = vs.get('Status') or 'Disabled'
                except Exception:
                    versioning = 'Disabled'
                
                # Logging
                try:
                    logging_enabled = 'Yes' if s3_bucket.get_bucket_logging(Bucket=name).get('LoggingEnabled') else 'No'
                except Exception:
                    logging_enabled = 'No'
                
                # Findings (same as before)
                if encrypted == 'No':
                    self._add_finding('HIGH', 'S3', name, 'No default encryption',
                        'Enable SSE-S3 or SSE-KMS on all buckets.')
                if public_access == 'Yes':
                    self._add_finding('CRITICAL', 'S3', name, 'Publicly accessible via ACL',
                        'Remove public ACLs; enable S3 Block Public Access.')
                if public_access == 'Partial block':
                    self._add_finding('HIGH', 'S3', name, 'Block Public Access not fully enabled',
                        'Enable all four S3 Block Public Access settings.')
                if versioning == 'Disabled':
                    self._add_finding('LOW', 'S3', name, 'Versioning not enabled',
                        'Enable versioning to protect against accidental deletion.')
                if logging_enabled == 'No':
                    self._add_finding('LOW', 'S3', name, 'Server access logging disabled',
                        'Enable S3 server access logging for audit trails.')
                
                data.append({
                    'Bucket Name':    name,
                    'Region':         bucket_region,
                    'Creation Date':  self._naive(bucket['CreationDate']),
                    'Public Access':  public_access,
                    'Encrypted':      encrypted,
                    'Versioning':     versioning,
                    'Access Logging': logging_enabled,
                })
                print("✓")
            except Exception as e:
                print(f"✗ FAILED: {e}")
                continue

        data.sort(key=lambda x: x['Bucket Name'])
        self._store('S3 Buckets', data)
        print(f"  → Completed: {len(data)} buckets processed")
        return len(data)
    #── 10. CloudTrail ────────────────────────────────────────────────────────
    def get_cloudtrail_events(self, days=7):
        print("Collecting CloudTrail events...")
        critical_actions = {
            'StartInstances','StopInstances','TerminateInstances','RunInstances','RebootInstances',
            'CreateVolume','DeleteVolume','AttachVolume','DetachVolume',
            'CreateSnapshot','DeleteSnapshot',
            'AuthorizeSecurityGroupIngress','RevokeSecurityGroupIngress',
            'CreateSecurityGroup','DeleteSecurityGroup',
            'CreateLoadBalancer','DeleteLoadBalancer',
            'CreateDBInstance','DeleteDBInstance','RebootDBInstance',
            'DeleteBucket','PutBucketPolicy',
            'CreateUser','DeleteUser','AttachUserPolicy','DetachUserPolicy',
            'CreateAccessKey','DeleteAccessKey',
        }
        trail_info = self._find_active_trail()
        if not trail_info:
            self._add_finding('CRITICAL', 'CloudTrail', 'Account', 'No active CloudTrail trail',
                'Enable CloudTrail in all regions immediately.')
            self._store('CloudTrail Events', [])
            return 0

        ct   = self._client('cloudtrail', trail_info['region'])
        try:
            trail    = ct.describe_trails(trailNameList=[trail_info['arn']])['trailList'][0]
            is_multi = trail.get('IsMultiRegionTrail', False)
            if not is_multi:
                self._add_finding('HIGH', 'CloudTrail', trail_info['name'],
                    'Trail is single-region only',
                    'Convert to multi-region trail to capture events from all regions.')
        except Exception:
            is_multi = False

        start = _now() - timedelta(days=days)
        end   = _now()

        def _fetch_action(action_region_pair):
            action, region = action_region_pair
            events = []
            try:
                client = self._client('cloudtrail', region)
                for event in client.lookup_events(
                    LookupAttributes=[{'AttributeKey': 'EventName', 'AttributeValue': action}],
                    StartTime=start, EndTime=end, MaxResults=50,
                ).get('Events', []):
                    username = event.get('Username', 'N/A')
                    try:
                        uid      = json.loads(event.get('CloudTrailEvent', '{}')).get('userIdentity', {})
                        username = uid.get('arn', uid.get('userName', username))
                    except Exception:
                        pass
                    resources     = event.get('Resources', [])
                    resource_name = resources[0].get('ResourceName', 'N/A') if resources else 'N/A'
                    events.append({
                        'Time':         self._naive(event.get('EventTime')),
                        'User':         username,
                        'Event Name':   event.get('EventName'),
                        'Resource':     resource_name,
                        'Region':       region,
                        'Source IP':    event.get('SourceIPAddress', 'N/A'),
                        'Event Source': event.get('EventSource', 'N/A'),
                    })
            except Exception:
                pass
            return events

        # If multi-region: query from one region; else query each region
        query_regions = [trail_info['region']] if is_multi else self.regions
        pairs         = [(a, r) for a in critical_actions for r in query_regions]

        ct_data = []
        with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(pairs))) as ex:
            for fut in as_completed([ex.submit(_fetch_action, p) for p in pairs]):
                ct_data.extend(fut.result())

        ct_data.sort(key=lambda x: x['Time'] or datetime.min, reverse=True)
        self._store('CloudTrail Events', ct_data)
        return len(ct_data)

    def _find_active_trail(self):
        for region in ['us-east-1', 'eu-west-1']:
            try:
                ct = self._client('cloudtrail', region)
                for trail in ct.describe_trails().get('trailList', []):
                    if ct.get_trail_status(Name=trail['TrailARN']).get('IsLogging'):
                        return {'region': region, 'arn': trail['TrailARN'], 'name': trail['Name']}
            except Exception:
                continue
        return None
    # ── 11. Cloudfront (parallel across regions) ───────────────────────
    def get_cloudfront(self):
        print("Collecting CloudFront distributions...")
        data = []
        try:
            cf = self._client('cloudfront', 'us-east-1')
            paginator = cf.get_paginator('list_distributions')
            for page in paginator.paginate():
                for dist in page['DistributionList'].get('Items', []):
                    # Check logging
                    logging_enabled = dist.get('Logging', {}).get('Enabled', False)
                    # Check WAF association
                    waf_web_acl_id = dist.get('WebACLId', '')
                    # Check default root object
                    default_root = dist.get('DefaultRootObject', '')
                    # Check TLS version
                    viewer_cert = dist.get('ViewerCertificate', {})
                    tls_version = viewer_cert.get('MinimumProtocolVersion', 'N/A')
                    data.append({
                        'Distribution ID': dist['Id'],
                        'Domain Name': dist['DomainName'],
                        'Status': dist['Status'],
                        'Enabled': dist['Enabled'],
                        'Logging Enabled': 'Yes' if logging_enabled else 'No',
                        'WAF Web ACL': waf_web_acl_id if waf_web_acl_id else 'None',
                        'Default Root Object': default_root if default_root else 'None',
                        'TLS Version': tls_version,
                        'Comment': dist.get('Comment', ''),
                    })
                    if not logging_enabled:
                        self._add_finding('MEDIUM', 'CloudFront', dist['Id'],
                                        'Access logging not enabled',
                                        'Enable CloudFront access logging for audit trail.')
                    if not waf_web_acl_id:
                        self._add_finding('MEDIUM', 'CloudFront', dist['Id'],
                                        'No WAF web ACL associated',
                                        'Associate a WAF web ACL to protect against common attacks.')
        except Exception as e:
            print(f"  Error CloudFront: {e}")
        self._store('CloudFront', data)
        return len(data)
    # ── 11. Aurora Clusters (parallel across regions) ───────────────────────
    def get_aurora_clusters(self):
        print("Collecting Aurora DB clusters...")
        def _region(region):
            rows = []
            try:
                rds = self._client('rds', region)
                clusters = rds.describe_db_clusters()['DBClusters']
                for cluster in clusters:
                    # Filter only Aurora engines
                    engine = cluster.get('Engine', '')
                    if 'aurora' not in engine.lower():
                        continue
                    enc = cluster.get('StorageEncrypted', False)
                    bkp = cluster.get('BackupRetentionPeriod', 0)
                    pub = cluster.get('PubliclyAccessible', False)
                    rows.append({
                        'Region': region,
                        'Cluster ID': cluster['DBClusterIdentifier'],
                        'Engine': engine,
                        'Engine Version': cluster.get('EngineVersion', 'N/A'),
                        'Status': cluster.get('Status', 'N/A'),
                        'Encrypted': 'Yes' if enc else 'No',
                        'Backup Retention (days)': bkp,
                        'Publicly Accessible': 'Yes' if pub else 'No',
                        'Multi-AZ': cluster.get('MultiAZ', 'No'),
                    })
                    if not enc:
                        self._add_finding('HIGH', 'Aurora', cluster['DBClusterIdentifier'],
                                        'Aurora cluster not encrypted at rest',
                                        'Enable storage encryption for Aurora cluster.')
                    if bkp < 7:
                        self._add_finding('MEDIUM', 'Aurora', cluster['DBClusterIdentifier'],
                                        f'Backup retention only {bkp} day(s)',
                                        'Set backup retention to at least 7 days.')
                    if pub:
                        self._add_finding('CRITICAL', 'Aurora', cluster['DBClusterIdentifier'],
                                        'Aurora cluster publicly accessible',
                                        'Disable public access and place in private subnet.')
            except Exception as e:
                print(f"    ⚠ Aurora error in {region}: {e}")
            return rows
        data = self._run_in_regions(_region)
        self._store('Aurora Clusters', data)
        return len(data)
    # ── 11. DynamoDB Tables (parallel across regions) ───────────────────────
    def get_dynamodb_tables(self):
        print("Collecting DynamoDB tables...")
        def _region(region):
            rows = []
            try:
                dynamodb = self._client('dynamodb', region)
                paginator = dynamodb.get_paginator('list_tables')
                for page in paginator.paginate():
                    for table_name in page['TableNames']:
                        desc = dynamodb.describe_table(TableName=table_name)['Table']
                        enc = desc.get('SSEDescription', {}).get('Status') == 'ENABLED'
                        pitr = False
                        try:
                            pitr_desc = dynamodb.describe_continuous_backups(TableName=table_name)
                            pitr = pitr_desc['ContinuousBackupsDescription']['PointInTimeRecoveryDescription']['PointInTimeRecoveryStatus'] == 'ENABLED'
                        except:
                            pass
                        rows.append({
                            'Region': region,
                            'Table Name': table_name,
                            'Status': desc.get('TableStatus', 'N/A'),
                            'Encrypted at Rest': 'Yes' if enc else 'No',
                            'Point-in-Time Recovery': 'Enabled' if pitr else 'Disabled',
                            'Item Count': desc.get('ItemCount', 0),
                            'Size (GB)': round(desc.get('TableSizeBytes', 0) / (1024**3), 2),
                        })
                        if not enc:
                            self._add_finding('HIGH', 'DynamoDB', table_name,
                                            'Table not encrypted at rest',
                                            'Enable default encryption (AWS owned or KMS).')
                        if not pitr:
                            self._add_finding('MEDIUM', 'DynamoDB', table_name,
                                            'Point-in-time recovery not enabled',
                                            'Enable PITR to protect against accidental writes/deletes.')
            except Exception as e:
                print(f"    ⚠ DynamoDB error in {region}: {e}")
            return rows
        data = self._run_in_regions(_region)
        self._store('DynamoDB Tables', data)
        return len(data)
    # ── 11. ElastiCache Clusters (parallel across regions) ───────────────────────
    def get_elasticache_clusters(self):
        print("Collecting ElastiCache clusters...")
        def _region(region):
            rows = []
            try:
                ec = self._client('elasticache', region)
                paginator = ec.get_paginator('describe_cache_clusters')
                for page in paginator.paginate(ShowCacheNodeInfo=False):
                    for cluster in page['CacheClusters']:
                        enc_at_rest = cluster.get('AtRestEncryptionEnabled', False)
                        enc_in_transit = cluster.get('TransitEncryptionEnabled', False)
                        backup_retention = cluster.get('SnapshotRetentionLimit', 0)
                        rows.append({
                            'Region': region,
                            'Cluster ID': cluster['CacheClusterId'],
                            'Engine': cluster['Engine'],
                            'Status': cluster['CacheClusterStatus'],
                            'Node Type': cluster.get('CacheNodeType', 'N/A'),
                            'Encryption at Rest': 'Yes' if enc_at_rest else 'No',
                            'Encryption in Transit': 'Yes' if enc_in_transit else 'No',
                            'Backup Retention (days)': backup_retention,
                        })
                        if not enc_at_rest:
                            self._add_finding('HIGH', 'ElastiCache', cluster['CacheClusterId'],
                                            'Encryption at rest not enabled',
                                            'Re-create cluster with at-rest encryption.')
                        if not enc_in_transit:
                            self._add_finding('HIGH', 'ElastiCache', cluster['CacheClusterId'],
                                            'Encryption in transit not enabled',
                                            'Enable transit encryption (requires Redis 6.x).')
                        if backup_retention == 0:
                            self._add_finding('MEDIUM', 'ElastiCache', cluster['CacheClusterId'],
                                            'Automated backups disabled',
                                            'Enable snapshots with retention ≥7 days.')
            except Exception as e:
                print(f"    ⚠ ElastiCache error in {region}: {e}")
            return rows
        data = self._run_in_regions(_region)
        self._store('ElastiCache Clusters', data)
        return len(data)
    # ── 11. CloudWatch Alarms (parallel across regions) ───────────────────────
    def get_cloudwatch_alarms(self):
        print("Collecting CloudWatch alarms...")
        def _region(region):
            rows = []
            cw   = self._client('cloudwatch', region)
            for page in cw.get_paginator('describe_alarms').paginate():
                for alarm in page['MetricAlarms']:
                    rows.append({
                        'Region':              region,
                        'Alarm Name':          alarm['AlarmName'],
                        'State':               alarm['StateValue'],
                        'State Updated':       self._naive(alarm['StateUpdatedTimestamp']),
                        'Metric Name':         alarm.get('MetricName', 'N/A'),
                        'Namespace':           alarm.get('Namespace', 'N/A'),
                        'Threshold':           alarm.get('Threshold', 'N/A'),
                        'Comparison Operator': alarm.get('ComparisonOperator', 'N/A'),
                    })
                    if alarm['StateValue'] == 'ALARM':
                        self._add_finding('HIGH', 'CloudWatch', alarm['AlarmName'],
                            f"Alarm in ALARM state ({alarm.get('MetricName','?')})",
                            'Investigate and resolve the underlying issue.')
            return rows

        data = self._run_in_regions(_region)
        if not data:
            self._add_finding('MEDIUM', 'CloudWatch', 'Account', 'No CloudWatch alarms configured',
                'Create alarms for CPU, disk, billing, error rates.')
        self._store('CloudWatch Alarms', data)
        return len(data)

    # ── 12. GuardDuty (parallel across regions) ───────────────────────────────
    def get_guardduty_status(self):
        print("Collecting GuardDuty status...")
        def _region(region):
            rows = []
            gd   = self._client('guardduty', region)
            dets = gd.list_detectors().get('DetectorIds', [])
            if not dets:
                rows.append({'Region': region, 'Status': 'NOT ENABLED', 'Detector ID': 'N/A', 'High Severity Findings': 0})
                self._add_finding('HIGH', 'GuardDuty', region, f'GuardDuty not enabled in {region}',
                    f'Enable GuardDuty in {region}.')
            else:
                for det_id in dets:
                    det   = gd.get_detector(DetectorId=det_id)
                    high  = len(gd.list_findings(
                        DetectorId=det_id,
                        FindingCriteria={'Criterion': {'severity': {'Gte': 7}}},
                    ).get('FindingIds', []))
                    rows.append({
                        'Region':                 region,
                        'Status':                 det.get('Status', 'N/A'),
                        'Detector ID':            det_id,
                        'High Severity Findings': high,
                    })
                    if high > 0:
                        self._add_finding('CRITICAL', 'GuardDuty', region,
                            f'{high} high-severity finding(s)',
                            'Review and remediate GuardDuty findings immediately.')
            return rows

        data = self._run_in_regions(_region)
        self._store('GuardDuty', data)
        return len(data)

    # ── 13. AWS Config (parallel across regions) ──────────────────────────────
    # (commented out in your original – left as is)

    # ── 14. VPC (parallel regions + parallel flow-log check per VPC) ────────────
    def get_vpc_info(self):
        print("Collecting VPC info...")
        def _region(region):
            ec2  = self._client('ec2', region)
            vpcs = ec2.describe_vpcs()['Vpcs']

            def _enrich(vpc):
                name       = next((t['Value'] for t in vpc.get('Tags', []) if t['Key'] == 'Name'), 'N/A')
                fl         = ec2.describe_flow_logs(
                    Filters=[{'Name': 'resource-id', 'Values': [vpc['VpcId']]}])['FlowLogs']
                flow_logs  = 'Enabled' if fl else 'Disabled'
                is_default = vpc.get('IsDefault', False)
                if is_default:
                    self._add_finding('LOW', 'VPC', vpc['VpcId'],
                        f'Default VPC exists in {region}', 'Delete default VPCs in unused regions.')
                if flow_logs == 'Disabled':
                    self._add_finding('MEDIUM', 'VPC', vpc['VpcId'],
                        f'VPC Flow Logs disabled ({region})',
                        'Enable VPC Flow Logs for network traffic visibility.')
                return {
                    'Region': region, 'VPC ID': vpc['VpcId'], 'Name': name,
                    'CIDR': vpc['CidrBlock'], 'Is Default': 'Yes' if is_default else 'No',
                    'State': vpc['State'], 'Flow Logs': flow_logs,
                }

            rows = []
            if vpcs:
                with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(vpcs))) as ex:
                    for fut in as_completed([ex.submit(_enrich, v) for v in vpcs]):
                        try:
                            rows.append(fut.result())
                        except Exception as e:
                            print(f"    ⚠ VPC enrich error: {e}")
            return rows

        data = self._run_in_regions(_region)
        self._store('VPC Info', data)
        return len(data)

    # ── 15. Load Balancers (parallel across regions) ──────────────────────────
    def get_load_balancers(self):
        print("Collecting load balancers...")
        def _region(region):
            rows  = []
            elbv2 = self._client('elbv2', region)
            for lb in elbv2.describe_load_balancers()['LoadBalancers']:
                tgs   = elbv2.describe_target_groups(LoadBalancerArn=lb['LoadBalancerArn'])['TargetGroups']
                attrs = {a['Key']: a['Value'] for a in
                         elbv2.describe_load_balancer_attributes(LoadBalancerArn=lb['LoadBalancerArn'])['Attributes']}
                acc_logs = attrs.get('access_logs.s3.enabled', 'false')
                del_prot = attrs.get('deletion_protection.enabled', 'false')
                rows.append({
                    'Region':              region,
                    'Type':                lb['Type'],
                    'Name':                lb['LoadBalancerName'],
                    'DNS Name':            lb['DNSName'],
                    'Scheme':              lb['Scheme'],
                    'Target Groups':       len(tgs),
                    'Access Logs':         acc_logs,
                    'Deletion Protection': del_prot,
                    'Created Time':        self._naive(lb['CreatedTime']),
                })
                if not tgs:
                    self._add_finding('LOW', 'Load Balancers', lb['LoadBalancerName'],
                        'No target groups (likely unused)', 'Remove unused load balancers.')
                if acc_logs == 'false':
                    self._add_finding('LOW', 'Load Balancers', lb['LoadBalancerName'],
                        'Access logging not enabled', 'Enable access logs on all load balancers.')
                if del_prot == 'false':
                    self._add_finding('LOW', 'Load Balancers', lb['LoadBalancerName'],
                        'Deletion protection not enabled',
                        'Enable deletion protection on production load balancers.')
            return rows

        data = self._run_in_regions(_region)
        self._store('Load Balancers', data)
        return len(data)

    # ── 16. Elastic IPs (parallel across regions) ─────────────────────────────
    def get_elastic_ips(self):
        print("Collecting Elastic IPs...")
        def _region(region):
            rows = []
            for addr in self._client('ec2', region).describe_addresses()['Addresses']:
                assoc = addr.get('AssociationId')
                rows.append({
                    'Region':               region,
                    'Allocation ID':        addr.get('AllocationId', 'N/A'),
                    'Public IP':            addr.get('PublicIp', 'N/A'),
                    'Status':               'Associated' if assoc else 'Unassociated',
                    'Instance ID':          addr.get('InstanceId', 'N/A'),
                    'Network Interface ID': addr.get('NetworkInterfaceId', 'N/A'),
                })
                if not assoc:
                    self._add_finding('LOW', 'Elastic IPs', addr.get('PublicIp', ''),
                        'Unassociated Elastic IP (idle cost)', 'Release unused Elastic IPs.')
            return rows

        data = self._run_in_regions(_region)
        self._store('Elastic IPs', data)
        return len(data)
    # ── 17. waf (global) ──────────────────────────────────────────────────
    def get_waf_status(self):
        print("Collecting AWS WAF (Regional & Global)...")
        data = []
        try:
            waf_global = self._client('wafv2', 'us-east-1')
            for scope in ['CLOUDFRONT', 'REGIONAL']:
                next_token = None
                while True:
                    params = {'Scope': scope}
                    if next_token:
                        params['NextMarker'] = next_token
                    response = waf_global.list_web_acls(**params)
                    for acl in response.get('WebACLs', []):
                        rules = []
                        for rule in acl.get('Rules', []):
                            action = list(rule.get('Action', {}).keys())[0] if rule.get('Action') else 'ALLOW'
                            rules.append(f"{rule['Name']}:{action}")
                        data.append({
                            'Scope': scope,
                            'ACL Name': acl['Name'],
                            'ACL ID': acl['Id'],
                            'Rules': ', '.join(rules),
                            'Capacity': acl.get('Capacity', 0),
                            'Default Action': list(acl.get('DefaultAction', {}).keys())[0] if acl.get('DefaultAction') else 'ALLOW',
                            'Description': acl.get('Description', ''),
                        })
                        if 'ALLOW' in rules and not any('BLOCK' in r for r in rules):
                            self._add_finding('MEDIUM', 'WAF', acl['Name'],
                                'WAF ACL has ALLOW default action but no BLOCK rules',
                                'Add rate‑based or threat‑detection rules to block malicious traffic.')
                    next_token = response.get('NextMarker')
                    if not next_token:
                        break
        except Exception as e:
            print(f"  Error collecting WAF: {e}")
            self._add_finding('LOW', 'WAF', 'Global', f'Error retrieving WAF data: {str(e)}', 'Check permissions and try again.')
        self._store('AWS WAF', data)
        return len(data)
    
    # ── 17. OpenSearch (Elasticsearch) (global) ──────────────────────────────────────────────────
    def get_opensearch_domains(self):
        print("Collecting OpenSearch domains...")
        def _region(region):
            rows = []
            try:
                os = self._client('opensearch', region)
                domains = os.list_domain_names()['DomainNames']
                for domain_info in domains:
                    domain = os.describe_domain(DomainName=domain_info['DomainName'])['DomainStatus']
                    enc = domain.get('EncryptionAtRestOptions', {}).get('Enabled', False)
                    logging = domain.get('LogPublishingOptions', {})
                    log_enabled = any(log.get('Enabled') for log in logging.values())
                    pub = domain.get('AccessPolicies', '')  # if contains '*' or open
                    is_public = '"Effect":"Allow"' in pub and '"Principal":"*"' in pub
                    rows.append({
                        'Region': region,
                        'Domain Name': domain['DomainName'],
                        'Engine': domain.get('EngineVersion', 'N/A'),
                        'Status': domain.get('Status', 'N/A'),
                        'Encrypted at Rest': 'Yes' if enc else 'No',
                        'Logging Enabled': 'Yes' if log_enabled else 'No',
                        'Publicly Accessible': 'Yes' if is_public else 'No',
                        'Instance Count': domain.get('ClusterConfig', {}).get('InstanceCount', 0),
                        'Instance Type': domain.get('ClusterConfig', {}).get('InstanceType', 'N/A'),
                    })
                    if not enc:
                        self._add_finding('HIGH', 'OpenSearch', domain['DomainName'], 'Encryption at rest disabled', 'Enable encryption.')
                    if not log_enabled:
                        self._add_finding('MEDIUM', 'OpenSearch', domain['DomainName'], 'Audit logging disabled', 'Enable log publishing.')
                    if is_public:
                        self._add_finding('CRITICAL', 'OpenSearch', domain['DomainName'], 'Publicly accessible domain', 'Restrict access with resource-based policy.')
            except Exception as e:
                print(f"    ⚠ OpenSearch error in {region}: {e}")
            return rows
        data = self._run_in_regions(_region)
        self._store('OpenSearch Domains', data)
        return len(data)
    # ── 17. Redshift  (Elasticsearch) (global) ──────────────────────────────────────────────────
    def get_redshift_clusters(self):
        print("Collecting Redshift clusters...")
        def _region(region):
            rows = []
            try:
                redshift = self._client('redshift', region)
                clusters = redshift.describe_clusters()['Clusters']
                for cluster in clusters:
                    enc = cluster.get('Encrypted', False)
                    logging = cluster.get('LoggingStatus', {}).get('LoggingEnabled', False)
                    pub = cluster.get('PubliclyAccessible', False)
                    bkp_retention = cluster.get('AutomatedSnapshotRetentionPeriod', 0)
                    rows.append({
                        'Region': region,
                        'Cluster ID': cluster['ClusterIdentifier'],
                        'Status': cluster['ClusterStatus'],
                        'Node Type': cluster['NodeType'],
                        'Encrypted': 'Yes' if enc else 'No',
                        'Logging Enabled': 'Yes' if logging else 'No',
                        'Publicly Accessible': 'Yes' if pub else 'No',
                        'Backup Retention (days)': bkp_retention,
                        'VPC ID': cluster.get('VpcId', 'N/A'),
                    })
                    if not enc:
                        self._add_finding('HIGH', 'Redshift', cluster['ClusterIdentifier'], 'Encryption disabled', 'Enable encryption at rest.')
                    if not logging:
                        self._add_finding('MEDIUM', 'Redshift', cluster['ClusterIdentifier'], 'Audit logging disabled', 'Enable logging to S3/CloudWatch.')
                    if pub:
                        self._add_finding('CRITICAL', 'Redshift', cluster['ClusterIdentifier'], 'Publicly accessible', 'Place in private subnet.')
                    if bkp_retention < 7:
                        self._add_finding('MEDIUM', 'Redshift', cluster['ClusterIdentifier'], f'Backup retention {bkp_retention} days', 'Increase to at least 7 days.')
            except Exception as e:
                print(f"    ⚠ Redshift error in {region}: {e}")
            return rows
        data = self._run_in_regions(_region)
        self._store('Redshift Clusters', data)
        return len(data)
    # ── 17. EKS   (Elasticsearch) (global) ──────────────────────────────────────────────────
    def get_eks_clusters(self):
        print("Collecting EKS clusters...")
        def _region(region):
            rows = []
            try:
                eks = self._client('eks', region)
                clusters = eks.list_clusters()['clusters']
                for cluster_name in clusters:
                    cluster = eks.describe_cluster(name=cluster_name)['cluster']
                    endpoint_public = cluster.get('resourcesVpcConfig', {}).get('endpointPublicAccess', False)
                    logging_enabled = any(cluster.get('logging', {}).get('clusterLogging', []))
                    rows.append({
                        'Region': region,
                        'Cluster Name': cluster['name'],
                        'Status': cluster['status'],
                        'Version': cluster['version'],
                        'Endpoint Public Access': 'Yes' if endpoint_public else 'No',
                        'Logging Enabled': 'Yes' if logging_enabled else 'No',
                        'VPC ID': cluster['resourcesVpcConfig'].get('vpcId', 'N/A'),
                        'Security Groups': ', '.join(cluster['resourcesVpcConfig'].get('securityGroupIds', [])),
                    })
                    if endpoint_public:
                        self._add_finding('HIGH', 'EKS', cluster['name'], 'Public endpoint access enabled', 'Disable public access or restrict with CIDR blocks.')
                    if not logging_enabled:
                        self._add_finding('MEDIUM', 'EKS', cluster['name'], 'Control plane logging disabled', 'Enable API/audit logs to CloudWatch.')
            except Exception as e:
                print(f"    ⚠ EKS error in {region}: {e}")
            return rows
        data = self._run_in_regions(_region)
        self._store('EKS Clusters', data)
        return len(data)
    # ── 17. ECS    (Elasticsearch) (global) ──────────────────────────────────────────────────
    def get_ecs_clusters(self):
        print("Collecting ECS clusters...")
        def _region(region):
            rows = []
            try:
                ecs = self._client('ecs', region)
                clusters = ecs.list_clusters()['clusterArns']
                for cluster_arn in clusters:
                    cluster = ecs.describe_clusters(clusters=[cluster_arn])['clusters'][0]
                    cluster_name = cluster['clusterName']
                    # Get services
                    services = ecs.list_services(cluster=cluster_name)['serviceArns']
                    # Check container insights
                    insights = cluster.get('settings', [])
                    insights_enabled = any(s['name'] == 'containerInsights' and s['value'] == 'enabled' for s in insights)
                    rows.append({
                        'Region': region,
                        'Cluster Name': cluster_name,
                        'Status': cluster['status'],
                        'Active Services': len(services),
                        'Running Tasks': cluster.get('runningTasksCount', 0),
                        'Pending Tasks': cluster.get('pendingTasksCount', 0),
                        'Container Insights': 'Enabled' if insights_enabled else 'Disabled',
                    })
                    if not insights_enabled:
                        self._add_finding('LOW', 'ECS', cluster_name, 'Container Insights disabled', 'Enable Container Insights for monitoring.')
            except Exception as e:
                print(f"    ⚠ ECS error in {region}: {e}")
            return rows
        data = self._run_in_regions(_region)
        self._store('ECS Clusters', data)
        return len(data)
    # ── 17. KMS     (Elasticsearch) (global) ──────────────────────────────────────────────────
    def get_kms_keys(self):
        print("Collecting KMS keys...")
        def _region(region):
            rows = []
            try:
                kms = self._client('kms', region)
                paginator = kms.get_paginator('list_keys')
                for page in paginator.paginate():
                    for key_info in page['Keys']:
                        key_id = key_info['KeyId']
                        try:
                            key_metadata = kms.describe_key(KeyId=key_id)['KeyMetadata']
                            rotation = kms.get_key_rotation_status(KeyId=key_id)
                            rotation_enabled = rotation.get('KeyRotationEnabled', False)
                            rows.append({
                                'Region': region,
                                'Key ID': key_id,
                                'Alias': key_metadata.get('AliasName', 'N/A'),
                                'Status': key_metadata.get('KeyState', 'N/A'),
                                'Origin': key_metadata.get('Origin', 'N/A'),
                                'Creation Date': self._naive(key_metadata.get('CreationDate')),
                                'Rotation Enabled': 'Yes' if rotation_enabled else 'No',
                                'Description': key_metadata.get('Description', ''),
                            })
                            if not rotation_enabled and key_metadata.get('KeySpec') != 'HMAC':
                                self._add_finding('MEDIUM', 'KMS', key_id, 'Key rotation not enabled', 'Enable automatic yearly rotation.')
                        except Exception as e:
                            print(f"      ⚠ Error describing key {key_id}: {e}")
            except Exception as e:
                print(f"    ⚠ KMS error in {region}: {e}")
            return rows
        data = self._run_in_regions(_region)
        self._store('KMS Keys', data)
        return len(data)   
    # ── 17. Route53 (global) ──────────────────────────────────────────────────
    def get_route53_zones(self):
        print("Collecting Route53 zones...")
        try:
            zones = self._client('route53', 'us-east-1').list_hosted_zones()['HostedZones']
            data  = [{
                'Zone Name':    z['Name'],
                'Zone ID':      z['Id'],
                'Type':         'Private' if z['Config']['PrivateZone'] else 'Public',
                'Record Count': z['ResourceRecordSetCount'],
                'Comment':      z['Config'].get('Comment', ''),
            } for z in zones]
            self._store('Route53 Zones', data)
            return len(data)
        except Exception as e:
            print(f"  Error Route53: {e}")
            self._store('Route53 Zones', [])
            return 0

    # ── 18. VPN (parallel across regions) ────────────────────────────────────
    def get_vpn_connections(self):
        print("Collecting VPN connections...")
        def _region(region):
            rows = []
            for vpn in self._client('ec2', region).describe_vpn_connections()['VpnConnections']:
                name    = next((t['Value'] for t in vpn.get('Tags', []) if t['Key'] == 'Name'), 'N/A')
                tunnels = vpn.get('VgwTelemetry', [])
                t1      = tunnels[0]['Status'] if tunnels else 'N/A'
                t2      = tunnels[1]['Status'] if len(tunnels) > 1 else 'N/A'
                rows.append({
                    'Region': region, 'VPN Name': name, 'VPN ID': vpn['VpnConnectionId'],
                    'State': vpn['State'], 'Type': vpn['Type'],
                    'Tunnel 1 Status': t1, 'Tunnel 2 Status': t2,
                    'Customer Gateway': vpn.get('CustomerGatewayId', 'N/A'),
                    'Virtual Private Gateway': vpn.get('VpnGatewayId', 'N/A'),
                })
                if 'DOWN' in (t1, t2):
                    self._add_finding('HIGH', 'VPN', name, f'Tunnel DOWN (T1:{t1} T2:{t2})',
                        'Investigate and restore VPN tunnel connectivity.')
            return rows

        data = self._run_in_regions(_region)
        self._store('VPN Connections', data)
        return len(data)

    # ── 19. AWS Backup (parallel across regions) ──────────────────────────────
    def get_aws_backup_status(self):
        print("Collecting AWS Backup jobs...")
        end_time   = _now()
        start_time = end_time - timedelta(days=7)

        def _region(region):
            rows   = []
            backup = self._client('backup', region)
            for page in backup.get_paginator('list_backup_jobs').paginate(
                ByCreatedBefore=end_time, ByCreatedAfter=start_time
            ):
                for job in page['BackupJobs']:
                    rows.append({
                        'Region':           region,
                        'Backup Job ID':    job.get('BackupJobId'),
                        'Resource Type':    job.get('ResourceType'),
                        'Resource ARN':     job.get('ResourceArn'),
                        'State':            job.get('State'),
                        'Creation Time':    self._naive(job.get('CreationDate')),
                        'Completion Time':  self._naive(job.get('CompletionDate')),
                        'Backup Size (GB)': round(job.get('BackupSizeInBytes', 0) / (1024**3), 2),
                    })
                    if job.get('State') == 'FAILED':
                        self._add_finding('HIGH', 'AWS Backup', job.get('ResourceArn', ''),
                            'Backup job failed', 'Investigate and fix the failed backup job.')
            return rows

        data = self._run_in_regions(_region)
        if not data:
            self._add_finding('HIGH', 'AWS Backup', 'Account',
                'No backup jobs found in last 7 days',
                'Configure AWS Backup plans for EC2, RDS, EBS, DynamoDB.')
        self._store('AWS Backup Jobs', data)

    # ── 20. Secrets Manager (global) ────────────────────────────────────────────

    def get_secrets_manager(self):
        print("Collecting Secrets Manager secrets...")
        def _region(region):
            rows = []
            try:
                sm = self._client('secretsmanager', region)
                paginator = sm.get_paginator('list_secrets')
                for page in paginator.paginate():
                    for secret in page['SecretList']:
                        name = secret['Name']
                        rotation_enabled = secret.get('RotationEnabled', False)
                        last_rotated = secret.get('LastRotatedDate')
                        rotation_days = secret.get('RotationRules', {}).get('AutomaticallyAfterDays', 'N/A')
                        days_since_rotation = ( _now() - last_rotated ).days if last_rotated and rotation_enabled else None

                        rows.append({
                            'Region': region,
                            'Secret Name': name,
                            'Rotation Enabled': 'Yes' if rotation_enabled else 'No',
                            'Last Rotated': self._naive(last_rotated) if last_rotated else 'Never',
                            'Rotation Period (days)': rotation_days,
                            'Days Since Rotation': days_since_rotation if days_since_rotation is not None else 'N/A',
                            'Tags': str(secret.get('Tags', [])),
                        })
                        if not rotation_enabled:
                            self._add_finding('MEDIUM', 'Secrets Manager', f"{name} ({region})",
                                            'Secret rotation not enabled',
                                            'Enable automatic rotation for secrets containing credentials.')
                        elif days_since_rotation and rotation_days != 'N/A' and days_since_rotation > int(rotation_days):
                            self._add_finding('LOW', 'Secrets Manager', f"{name} ({region})",
                                            f'Secret last rotated {days_since_rotation} days ago (exceeds {rotation_days} day period)',
                                            'Rotate the secret manually or check rotation Lambda.')
            except Exception as e:
                print(f"    ⚠ Secrets Manager error in {region}: {e}")
            return rows
        data = self._run_in_regions(_region)
        self._store('Secrets Manager', data)
        return len(data)

    # ── 20. Cost Overview (global) ────────────────────────────────────────────
    def get_cost_overview(self, days=30):
        """
        Collects daily cost per service and performs analysis.
        Stores raw data + 3 analysis tables in self.report_data.
        """
        print("Collecting and analysing cost overview...")
        raw_data = []
        try:
            ce = self._client('ce', 'us-east-1')
            end_date = _now().date()
            start_date = end_date - timedelta(days=days)

            start = start_date.strftime('%Y-%m-%d')
            end = (end_date + timedelta(days=1)).strftime('%Y-%m-%d')

            next_token = None
            while True:
                params = {
                    'TimePeriod': {'Start': start, 'End': end},
                    'Granularity': 'DAILY',
                    'Metrics': ['UnblendedCost'],
                    'GroupBy': [{'Type': 'DIMENSION', 'Key': 'SERVICE'}]
                }
                if next_token:
                    params['NextPageToken'] = next_token

                response = ce.get_cost_and_usage(**params)

                for period in response['ResultsByTime']:
                    date = period['TimePeriod']['Start']
                    for group in period['Groups']:
                        service = group['Keys'][0]
                        amount = float(group['Metrics']['UnblendedCost']['Amount'])
                        if amount > 0.01:
                            raw_data.append({
                                'Date': date,
                                'Service': service,
                                'Cost (USD)': round(amount, 4)
                            })

                next_token = response.get('NextPageToken')
                if not next_token:
                    break

            if not raw_data:
                self._add_finding('LOW', 'Cost', 'Account',
                                'No cost data retrieved (check permissions / enable Cost Explorer)',
                                'Enable Cost Explorer and grant "ce:GetCostAndUsage".')
                self._store('Cost Overview (raw)', [])
                self._store('Cost Summary', [])
                self._store('Daily Total Cost', [])
                self._store('Cost Anomalies', [])
                return 0

            # ---- ANALYSIS ----
            df = pd.DataFrame(raw_data)
            df['Date'] = pd.to_datetime(df['Date'])
            df = df.sort_values('Date')

            # 1. Cost Summary (per service)
            summary = df.groupby('Service')['Cost (USD)'].agg(['sum', 'mean', 'count']).reset_index()
            summary.columns = ['Service', 'Total (USD)', 'Average Daily (USD)', 'Days with Cost']
            total_all = summary['Total (USD)'].sum()
            summary['% of Total'] = (summary['Total (USD)'] / total_all * 100).round(2)
            summary = summary.sort_values('Total (USD)', ascending=False)

            # 2. Daily Total Cost
            daily = df.groupby('Date')['Cost (USD)'].sum().reset_index()
            daily.columns = ['Date', 'Total Cost (USD)']
            daily['Day-over-Day % Change'] = daily['Total Cost (USD)'].pct_change() * 100
            daily['7-Day Moving Avg'] = daily['Total Cost (USD)'].rolling(window=7, min_periods=1).mean()

            # 3. Anomaly detection
            mean_cost = daily['Total Cost (USD)'].mean()
            std_cost = daily['Total Cost (USD)'].std()
            anomalies = daily[
                (daily['Total Cost (USD)'] > mean_cost + 2*std_cost) |
                (daily['Total Cost (USD)'] < mean_cost - 2*std_cost)
            ].copy()
            if not anomalies.empty:
                anomalies['Deviation (USD)'] = anomalies['Total Cost (USD)'] - mean_cost
                anomalies['% Deviation'] = (anomalies['Deviation (USD)'] / mean_cost * 100).round(1)

            # Store all four
            self._store('Cost Overview (raw)', raw_data)
            self._store('Cost Summary', summary.to_dict('records'))
            self._store('Daily Total Cost', daily.to_dict('records'))
            self._store('Cost Anomalies', anomalies.to_dict('records') if not anomalies.empty else [])
            

            return len(raw_data)

        except Exception as e:
            print(f"  Error in Cost Explorer: {e}")
            self._add_finding('MEDIUM', 'Cost', 'Account',
                            f'Cost Explorer error: {str(e)}',
                            'Ensure Cost Explorer is enabled and the IAM role has "ce:GetCostAndUsage".')
            self._store('Cost Overview (raw)', [{'Note': str(e)}])
            return 0

    # ── Report Generation ───────────────────────────────────────────────────────

    # ─────────────────────────────────────────────────────────────────────────
    # ACCOUNT OVERVIEW
    # ─────────────────────────────────────────────────────────────────────────
    def build_account_overview(self):
        rd = self.report_data
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M UTC')

        # Severity counts
        sev = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0, 'INFO': 0}
        for f in self.findings:
            sev[f['Severity']] = sev.get(f['Severity'], 0) + 1

        def row(cat, metric, value, rec=''):
            if value == 0 or value == '0' or (isinstance(value, (int, float)) and value == 0):
                rec = ''
            return {'Category': cat, 'Metric': metric, 'Value': value, 'Recommendation': rec}
        def sep(title):
            return {'Category': f'── {title} ──', 'Metric': '', 'Value': '', 'Recommendation': ''}
        def blank():
            return {'Category': '', 'Metric': '', 'Value': '', 'Recommendation': ''}
        def add_section(title, rows):
            nonlocal overview
            if any(r['Value'] not in ('', 0, '0', 'N/A', 'Never') for r in rows if isinstance(r['Value'], (int, float, str))):
                overview.append(sep(title))
                overview.extend(rows)
                overview.append(blank())

        overview = []

        # ACCOUNT header
        overview.append(sep('ACCOUNT'))
        overview.append(row('Account', 'Account ID', self.account_id, ''))
        overview.append(row('Account', 'Account Name', self._account_name, ''))
        overview.append(row('Account', 'Report Generated', now_str, ''))
        overview.append(row('Account', 'Regions Scanned', len(self.regions), ''))
        overview.append(blank())

        # FINDINGS SUMMARY
        overview.append(sep('FINDINGS SUMMARY'))
        overview.append(row('Findings', '🔴 CRITICAL', sev['CRITICAL'], ''))
        overview.append(row('Findings', '🟠 HIGH', sev['HIGH'], ''))
        overview.append(row('Findings', '🟡 MEDIUM', sev['MEDIUM'], ''))
        overview.append(row('Findings', '🟢 LOW', sev['LOW'], ''))
        overview.append(row('Findings', '🔵 INFO', sev['INFO'], ''))
        overview.append(row('Findings', 'Total', len(self.findings), ''))
        overview.append(blank())

        # ---- COMPUTE ----
        ec2 = rd.get('EC2 Instances', [])
        ec2_total = len(ec2)
        ec2_running = sum(1 for i in ec2 if i.get('State') == 'running')
        ec2_stopped = sum(1 for i in ec2 if i.get('State') == 'stopped')
        ec2_public = sum(1 for i in ec2 if i.get('Public IP') not in ('N/A', None, ''))
        ec2_imdsv1 = sum(1 for i in ec2 if i.get('IMDSv2 Only') == 'No')
        compute_rows = [
            row('EC2', 'Total Instances', ec2_total, ''),
            row('EC2', 'Running', ec2_running, ''),
            row('EC2', 'Stopped', ec2_stopped, 'Terminate if not needed'),
            row('EC2', 'With Public IP', ec2_public, 'Restrict access or move to private subnet'),
            row('EC2', 'IMDSv1 Enabled (risk)', ec2_imdsv1, 'Enforce IMDSv2'),
        ]
        add_section('COMPUTE', compute_rows)

        # ---- STORAGE (EBS, Snapshots, AMIs, S3) ----
        ebs = rd.get('EBS Volumes', [])
        ebs_total = len(ebs)
        ebs_unattached = sum(1 for v in ebs if v.get('Attached To') == 'Unattached')
        ebs_unencrypted = sum(1 for v in ebs if v.get('Encrypted') == 'No')
        snapshots = rd.get('EBS Snapshots', [])
        snap_total = len(snapshots)
        snap_old = sum(1 for s in snapshots if s.get('Status') == 'Old (>90d)')
        amis = rd.get('AMIs', [])
        ami_total = len(amis)
        ami_old = sum(1 for a in amis if isinstance(a.get('Age (days)'), int) and a.get('Age (days)', 0) > 90)
        s3 = rd.get('S3 Buckets', [])
        s3_total = len(s3)
        s3_public = sum(1 for b in s3 if b.get('Public Access') == 'Yes')
        s3_unencrypted = sum(1 for b in s3 if b.get('Encrypted') == 'No')
        s3_no_versioning = sum(1 for b in s3 if b.get('Versioning') != 'Enabled')
        storage_rows = [
            row('EBS', 'Total Volumes', ebs_total, ''),
            row('EBS', 'Unattached', ebs_unattached, 'Delete or snapshot'),
            row('EBS', 'Unencrypted', ebs_unencrypted, 'Enable encryption'),
            row('EBS Snapshots', 'Total Snapshots', snap_total, ''),
            row('EBS Snapshots', 'Old (>90 days)', snap_old, 'Delete old snapshots'),
            row('AMIs', 'Total AMIs', ami_total, ''),
            row('AMIs', 'Old (>90 days)', ami_old, 'Deregister unused AMIs'),
            row('S3', 'Total Buckets', s3_total, ''),
            row('S3', 'Public', s3_public, 'Block public access'),
            row('S3', 'Unencrypted', s3_unencrypted, 'Enable default encryption'),
            row('S3', 'No Versioning', s3_no_versioning, 'Enable versioning'),
        ]
        add_section('STORAGE', storage_rows)

        # ---- DATABASE SERVICES (RDS, Aurora, DynamoDB, ElastiCache) ----
        rds = rd.get('RDS Instances', [])
        rds_total = len(rds)
        rds_public = sum(1 for r in rds if r.get('Publicly Accessible') == 'Yes')
        rds_unencrypted = sum(1 for r in rds if r.get('Encrypted') == 'No')
        rds_no_multi_az = sum(1 for r in rds if r.get('Multi-AZ') == 'No')
        db_rows = []
        if rds_total > 0:
            db_rows.append(row('RDS', 'Total Instances', rds_total, ''))
        if rds_public > 0:
            db_rows.append(row('RDS', 'Publicly Accessible', rds_public, 'Move to private subnet'))
        if rds_unencrypted > 0:
            db_rows.append(row('RDS', 'Unencrypted', rds_unencrypted, 'Enable encryption'))
        if rds_no_multi_az > 0:
            db_rows.append(row('RDS', 'No Multi-AZ', rds_no_multi_az, 'Enable Multi-AZ for production'))

        # Aurora
        aurora = rd.get('Aurora Clusters', [])
        aurora_total = len(aurora)
        aurora_unencrypted = sum(1 for c in aurora if c.get('Encrypted') == 'No')
        aurora_public = sum(1 for c in aurora if c.get('Publicly Accessible') == 'Yes')
        aurora_bkp_low = sum(1 for c in aurora if isinstance(c.get('Backup Retention (days)'), int) and c.get('Backup Retention (days)', 0) < 7)
        if aurora_total > 0:
            db_rows.append(row('Aurora', 'Total Clusters', aurora_total, ''))
        if aurora_unencrypted > 0:
            db_rows.append(row('Aurora', 'Unencrypted', aurora_unencrypted, 'Enable storage encryption'))
        if aurora_public > 0:
            db_rows.append(row('Aurora', 'Publicly Accessible', aurora_public, 'Move to private subnet'))
        if aurora_bkp_low > 0:
            db_rows.append(row('Aurora', 'Backup <7 days', aurora_bkp_low, 'Increase backup retention'))

        # DynamoDB
        dynamodb = rd.get('DynamoDB Tables', [])
        ddb_total = len(dynamodb)
        ddb_unencrypted = sum(1 for t in dynamodb if t.get('Encrypted at Rest') == 'No')
        ddb_no_pitr = sum(1 for t in dynamodb if t.get('Point-in-Time Recovery') != 'Enabled')
        if ddb_total > 0:
            db_rows.append(row('DynamoDB', 'Total Tables', ddb_total, ''))
        if ddb_unencrypted > 0:
            db_rows.append(row('DynamoDB', 'Unencrypted', ddb_unencrypted, 'Enable default encryption'))
        if ddb_no_pitr > 0:
            db_rows.append(row('DynamoDB', 'PITR Disabled', ddb_no_pitr, 'Enable point-in-time recovery'))

        # ElastiCache
        elasticache = rd.get('ElastiCache Clusters', [])
        ec_total = len(elasticache)
        ec_enc_at_rest = sum(1 for c in elasticache if c.get('Encryption at Rest') == 'No')
        ec_enc_transit = sum(1 for c in elasticache if c.get('Encryption in Transit') == 'No')
        ec_no_backup = sum(1 for c in elasticache if c.get('Backup Retention (days)', 0) == 0)
        if ec_total > 0:
            db_rows.append(row('ElastiCache', 'Total Clusters', ec_total, ''))
        if ec_enc_at_rest > 0:
            db_rows.append(row('ElastiCache', 'No Encrypt at Rest', ec_enc_at_rest, 'Enable at-rest encryption'))
        if ec_enc_transit > 0:
            db_rows.append(row('ElastiCache', 'No Encrypt in Transit', ec_enc_transit, 'Enable in-transit encryption'))
        if ec_no_backup > 0:
            db_rows.append(row('ElastiCache', 'Backups Disabled', ec_no_backup, 'Enable snapshots with retention'))

        if db_rows:
            add_section('DATABASE SERVICES', db_rows)

        # ---- BIG DATA & ANALYTICS (Redshift, OpenSearch) ----
        analytics_rows = []
        redshift = rd.get('Redshift Clusters', [])
        rs_total = len(redshift)
        rs_unencrypted = sum(1 for c in redshift if c.get('Encrypted') == 'No')
        rs_public = sum(1 for c in redshift if c.get('Publicly Accessible') == 'Yes')
        rs_no_logging = sum(1 for c in redshift if c.get('Logging Enabled') == 'No')
        rs_bkp_low = sum(1 for c in redshift if c.get('Backup Retention (days)', 0) < 7)
        if rs_total > 0:
            analytics_rows.append(row('Redshift', 'Total Clusters', rs_total, ''))
        if rs_unencrypted > 0:
            analytics_rows.append(row('Redshift', 'Unencrypted', rs_unencrypted, 'Enable encryption at rest'))
        if rs_public > 0:
            analytics_rows.append(row('Redshift', 'Publicly Accessible', rs_public, 'Move to private subnet'))
        if rs_no_logging > 0:
            analytics_rows.append(row('Redshift', 'Logging Disabled', rs_no_logging, 'Enable audit logging'))
        if rs_bkp_low > 0:
            analytics_rows.append(row('Redshift', 'Backup <7 days', rs_bkp_low, 'Increase backup retention'))

        opensearch = rd.get('OpenSearch Domains', [])
        os_total = len(opensearch)
        os_unencrypted = sum(1 for d in opensearch if d.get('Encrypted at Rest') == 'No')
        os_public = sum(1 for d in opensearch if d.get('Publicly Accessible') == 'Yes')
        os_no_logging = sum(1 for d in opensearch if d.get('Logging Enabled') == 'No')
        if os_total > 0:
            analytics_rows.append(row('OpenSearch', 'Total Domains', os_total, ''))
        if os_unencrypted > 0:
            analytics_rows.append(row('OpenSearch', 'Unencrypted', os_unencrypted, 'Enable encryption at rest'))
        if os_public > 0:
            analytics_rows.append(row('OpenSearch', 'Publicly Accessible', os_public, 'Restrict access policy'))
        if os_no_logging > 0:
            analytics_rows.append(row('OpenSearch', 'Logging Disabled', os_no_logging, 'Enable audit logging'))

        if analytics_rows:
            add_section('BIG DATA & ANALYTICS', analytics_rows)

        # ---- CONTAINER SERVICES (EKS, ECS) ----
        container_rows = []
        eks = rd.get('EKS Clusters', [])
        eks_total = len(eks)
        eks_public = sum(1 for c in eks if c.get('Endpoint Public Access') == 'Yes')
        eks_no_logging = sum(1 for c in eks if c.get('Logging Enabled') == 'No')
        if eks_total > 0:
            container_rows.append(row('EKS', 'Total Clusters', eks_total, ''))
        if eks_public > 0:
            container_rows.append(row('EKS', 'Public Endpoint', eks_public, 'Disable or restrict with CIDR'))
        if eks_no_logging > 0:
            container_rows.append(row('EKS', 'Control Plane Logging Disabled', eks_no_logging, 'Enable API/audit logs'))

        ecs = rd.get('ECS Clusters', [])
        ecs_total = len(ecs)
        ecs_no_insights = sum(1 for c in ecs if c.get('Container Insights') == 'Disabled')
        if ecs_total > 0:
            container_rows.append(row('ECS', 'Total Clusters', ecs_total, ''))
        if ecs_no_insights > 0:
            container_rows.append(row('ECS', 'Container Insights Disabled', ecs_no_insights, 'Enable for monitoring'))

        if container_rows:
            add_section('CONTAINER SERVICES', container_rows)

        # ---- CONTENT DELIVERY (CloudFront) ----
        cf = rd.get('CloudFront', [])
        cf_total = len(cf)
        cf_enabled = sum(1 for d in cf if d.get('Enabled') == True)
        cf_disabled = sum(1 for d in cf if d.get('Enabled') == False)
        cf_no_logging = sum(1 for d in cf if d.get('Logging Enabled') == 'No')
        cf_no_waf = sum(1 for d in cf if d.get('WAF Web ACL') in (None, '', 'None'))
        cf_rows = []
        if cf_total > 0:
            cf_rows.append(row('CloudFront', 'Total Distributions', cf_total, ''))
            cf_rows.append(row('CloudFront', 'Enabled Distributions', cf_enabled, ''))
            cf_rows.append(row('CloudFront', 'Disabled Distributions', cf_disabled, 'Review or delete'))
            if cf_no_logging > 0:
                cf_rows.append(row('CloudFront', 'Logging Disabled', cf_no_logging, 'Enable access logging'))
            if cf_no_waf > 0:
                cf_rows.append(row('CloudFront', 'No WAF ACL', cf_no_waf, 'Associate WAF web ACL'))
        if cf_rows:
            add_section('CONTENT DELIVERY', cf_rows)

        # ---- SECURITY SERVICES (WAF, Secrets Manager, KMS) ----
        sec_rows = []
        waf = rd.get('AWS WAF', [])
        waf_total = len(waf)
        waf_allow_default = sum(1 for acl in waf if acl.get('Default Action') == 'ALLOW' and 'BLOCK' not in acl.get('Rules', ''))
        if waf_total > 0:
            sec_rows.append(row('WAF', 'Total Web ACLs', waf_total, ''))
            if waf_allow_default > 0:
                sec_rows.append(row('WAF', 'Allow Default + No Block Rules', waf_allow_default, 'Add rate‑based or threat rules'))

        secrets = rd.get('Secrets Manager', [])
        secrets_total = len(secrets)
        secrets_no_rotation = sum(1 for s in secrets if s.get('Rotation Enabled') == 'No')
        secrets_overdue = sum(1 for s in secrets if isinstance(s.get('Days Since Rotation'), int) and s.get('Days Since Rotation', 0) > 90)
        if secrets_total > 0:
            sec_rows.append(row('Secrets Manager', 'Total Secrets', secrets_total, ''))
            if secrets_no_rotation > 0:
                sec_rows.append(row('Secrets Manager', 'Rotation Disabled', secrets_no_rotation, 'Enable auto‑rotation'))
            if secrets_overdue > 0:
                sec_rows.append(row('Secrets Manager', 'Overdue Rotation', secrets_overdue, 'Rotate manually or check Lambda'))

        kms = rd.get('KMS Keys', [])
        kms_total = len(kms)
        kms_rotation_disabled = sum(1 for k in kms if k.get('Rotation Enabled') == 'No')
        if kms_total > 0:
            sec_rows.append(row('KMS', 'Total Keys', kms_total, ''))
        if kms_rotation_disabled > 0:
            sec_rows.append(row('KMS', 'Rotation Disabled', kms_rotation_disabled, 'Enable automatic yearly rotation'))

        if sec_rows:
            add_section('SECURITY SERVICES', sec_rows)

        # ---- IDENTITY (IAM) ----
        iam = rd.get('IAM Users', [])
        iam_total = len(iam)
        iam_active = sum(1 for u in iam if u.get('Status') == 'Active')
        iam_inactive = sum(1 for u in iam if u.get('Status') in ('Inactive','Never Active'))
        iam_no_mfa = sum(1 for u in iam if u.get('MFA Enabled') == 'No')
        iam_old_keys = sum(1 for u in iam if u.get('Old Keys (>90d)', 0) > 0)
        # IAM Password Policy existence
        try:
            iam_client = self._client('iam', 'us-east-1')
            iam_client.get_account_password_policy()
            password_policy_exists = 'Yes'
            pwd_rec = ''
        except iam_client.exceptions.NoSuchEntityException:
            password_policy_exists = 'No'
            pwd_rec = 'Create a strong password policy (min length 14, require symbols, expiry ≤90 days)'
        except Exception:
            password_policy_exists = 'Error'
            pwd_rec = 'Check IAM permissions'
        identity_rows = [
            row('IAM', 'Total Users', iam_total, ''),
            row('IAM', 'Active', iam_active, ''),
            row('IAM', 'Inactive / Never', iam_inactive, 'Remove or disable'),
            row('IAM', 'Without MFA', iam_no_mfa, 'Enable MFA'),
            row('IAM', 'Old Access Keys', iam_old_keys, 'Rotate keys'),
            row('IAM', 'Password Policy Exists', password_policy_exists, pwd_rec),
        ]
        add_section('IDENTITY', identity_rows)

        # ---- NETWORKING (Security Groups + Elastic IPs) ----
        security_groups = rd.get('Security Groups', [])
        sg_total = len(security_groups)
        sg_open = sum(1 for sg in security_groups if sg.get('Open to 0.0.0.0/0') == 'Yes')
        eips = rd.get('Elastic IPs', [])
        eip_total = len(eips)
        eip_unassoc = sum(1 for e in eips if e.get('Status') == 'Unassociated')
        networking_rows = []
        networking_rows.append(row('Security Groups', 'Total', sg_total, ''))
        networking_rows.append(row('Security Groups', 'Open to 0.0.0.0/0', sg_open, 'Restrict ingress to known IPs'))
        if eip_total > 0:
            networking_rows.append(row('Elastic IPs', 'Total', eip_total, ''))
        if eip_unassoc > 0:
            networking_rows.append(row('Elastic IPs', 'Unassociated', eip_unassoc, 'Release unused EIPs'))
        add_section('NETWORKING', networking_rows)

        # ---- MONITORING & SECURITY (CloudWatch, GuardDuty) ----
        alarms = rd.get('CloudWatch Alarms', [])
        alarm_total = len(alarms)
        alarm_alarm = sum(1 for a in alarms if a.get('State') == 'ALARM')
        gd = rd.get('GuardDuty', [])
        gd_enabled = sum(1 for g in gd if g.get('Status') == 'ENABLED')
        gd_not_enabled = sum(1 for g in gd if g.get('Status') == 'NOT ENABLED')
        gd_high = sum(g.get('High Severity Findings', 0) for g in gd if isinstance(g.get('High Severity Findings'), int))
        mon_rows = [
            row('CloudWatch', 'Total Alarms', alarm_total, ''),
            row('CloudWatch', 'In ALARM state', alarm_alarm, 'Investigate'),
            row('GuardDuty', 'Regions Enabled', gd_enabled, ''),
            row('GuardDuty', 'NOT Enabled', gd_not_enabled, 'Enable GuardDuty'),
            row('GuardDuty', 'High Findings', gd_high, 'Remediate immediately'),
        ]
        add_section('MONITORING & SECURITY', mon_rows)

        # ---- BACKUP ----
        bkp = rd.get('AWS Backup Jobs', [])
        bkp_total = len(bkp)
        bkp_failed = sum(1 for b in bkp if b.get('State') == 'FAILED')
        bkp_completed = sum(1 for b in bkp if b.get('State') == 'COMPLETED')
        backup_rows = []
        backup_rows.append(row('AWS Backup', 'Jobs (last 7d)', bkp_total, ''))
        backup_rows.append(row('AWS Backup', 'Failed', bkp_failed, 'Investigate if >0'))
        if bkp_completed > 0:
            backup_rows.append(row('AWS Backup', 'Completed', bkp_completed, ''))
        add_section('BACKUP', backup_rows)

        self._store('Account Overview', overview)

    # ─────────────────────────────────────────────────────────────────────────
    # EXCEL
    # ─────────────────────────────────────────────────────────────────────────
    def _clean_for_excel(self, data):
        cleaned = []
        for item in data:
            row = {}
            for k, v in item.items():
                if isinstance(v, datetime):
                    row[k] = self._naive(v)
                elif isinstance(v, (dict, list)):
                    row[k] = str(v)
                else:
                    row[k] = v
            cleaned.append(row)
        return cleaned

    def generate_excel_bytes(self):
        print("Generating Excel report with action‑required row coloring & snapshot cell coloring...")
        output = io.BytesIO()
        priority = ['Account Overview', 'Recommendations & Gaps']

        # Desired order after priority sheets
        desired_order = [
            'IAM Users', 
            'AMIs',
            'EBS Volumes', 'EBS Snapshots',
            'EC2 Instances',
            'Security Groups',
            'Elastic IPs',
            'AWS WAF',
            'RDS Instances', 'Aurora Clusters',
            'DynamoDB Tables', 'ElastiCache Clusters',
            'AWS Backup Jobs',            
            'S3 Buckets',
            'Secrets Manager',
            'KMS Keys',
            'OpenSearch Domains',
            'Redshift Clusters',
            'EKS Clusters',
            'ECS Clusters',
            'CloudFront',
            'VPC Info',
            'Load Balancers',
            'Route53 Zones',
            'VPN Connections',
            'CloudWatch Alarms',
            'GuardDuty',
            'Secrets Manager',
            'CloudTrail Events',
            'Cost Dashboard',
            'IAM Roles', 'IAM Password Policy',
            # 'Monthly Cost by Service',
            'Cost Summary',
            'Daily Total Cost',
            'Cost Anomalies',
        ]

        # Build final sheet list: priority sheets first, then desired_order that exist, then any remaining
        all_sheets = list(self.report_data.keys())
        other_sheets = []
        for sheet in desired_order:
            if sheet in all_sheets and sheet not in priority:
                other_sheets.append(sheet)
        # Append any remaining sheets not in priority or desired_order
        for sheet in all_sheets:
            if sheet not in priority and sheet not in other_sheets:
                other_sheets.append(sheet)

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Write all sheets as before
            ov = self._clean_for_excel(self.report_data.get('Account Overview', []))
            if ov:
                pd.DataFrame(ov).to_excel(writer, sheet_name='Account Overview', index=False)

            sev_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3, 'INFO': 4}
            sf = sorted(self.findings, key=lambda x: sev_order.get(x['Severity'], 9))
            if sf:
                pd.DataFrame(sf).to_excel(writer, sheet_name='Recommendations & Gaps', index=False)

            for sheet in other_sheets:
                raw = self.report_data.get(sheet, [])
                if not isinstance(raw, list) or not raw:
                    continue
                if sheet == 'Cost Overview (raw)':
                    continue
                safe_name = sheet[:31].replace('\\', '_').replace('/', '_').replace('*', '_').replace('?', '_').replace(':', '_')
                pd.DataFrame(self._clean_for_excel(raw)).to_excel(writer, sheet_name=safe_name, index=False)

            # ========== STYLING & CONDITIONAL COLORING ==========
            wb = writer.book
            hdr_fill = PatternFill('solid', fgColor='1F4E79')
            hdr_font = Font(color='FFFFFF', bold=True)
            section_font = Font(bold=True, color='1F4E79')
            section_fill = PatternFill('solid', fgColor='D6EAF8')
            action_fill = PatternFill('solid', fgColor='FFCCCC')          # light red
            snapshot_old_fill = PatternFill('solid', fgColor='FFD9B3')    # orange
            snapshot_recent_fill = PatternFill('solid', fgColor='D9F0B3') # light green

            for ws in wb.worksheets:
                # Auto-width columns
                for col in ws.columns:
                    max_len = max((len(str(c.value or '')) for c in col), default=0)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)

                # Header row styling
                for cell in ws[1]:
                    cell.fill = hdr_fill
                    cell.font = hdr_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.row_dimensions[1].height = 22
                ws.freeze_panes = 'A2'

                # ----- 1. Recommendations & Gaps colouring (full row) -----
                if ws.title == 'Recommendations & Gaps':
                    for row in ws.iter_rows(min_row=2):
                        sev = str(row[0].value or '')
                        if sev in COLOUR:
                            fill = PatternFill('solid', fgColor=COLOUR[sev])
                            for cell in row:
                                cell.fill = fill

                # ----- 2. Account Overview styling (section headers + value cells) -----
                if ws.title == 'Account Overview':
                    # Define border styles
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    thick_border = Border(
                        left=Side(style='thick'), right=Side(style='thick'),
                        top=Side(style='thick'), bottom=Side(style='thick')
                    )

                    # First pass: apply section header styling and merge headers
                    section_start_rows = []
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                        if str(row[0].value or '').startswith('──'):
                            # Merge cells A:D for this section header
                            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
                            # Apply font and fill
                            for cell in row:
                                cell.font = section_font
                                cell.fill = section_fill
                            section_start_rows.append(row_idx)

                    # Add a sentinel for the end of the sheet
                    max_row = ws.max_row
                    section_start_rows.append(max_row + 1)

                    # Apply borders to each section
                    for i in range(len(section_start_rows) - 1):
                        start_row = section_start_rows[i]
                        end_row = section_start_rows[i + 1] - 1
                        if end_row < start_row:
                            continue

                        # Thick border for the whole section (outer box)
                        for col in range(1, 5):  # columns A-D
                            # Top border of first row of section (header)
                            top_cell = ws.cell(row=start_row, column=col)
                            top_cell.border = Border(
                                left=thin_border.left if col > 1 else thick_border.left,
                                right=thin_border.right if col < 4 else thick_border.right,
                                top=thick_border.top,
                                bottom=thin_border.bottom
                            )
                            # Bottom border of last row of section
                            bottom_cell = ws.cell(row=end_row, column=col)
                            bottom_cell.border = Border(
                                left=thin_border.left if col > 1 else thick_border.left,
                                right=thin_border.right if col < 4 else thick_border.right,
                                top=thin_border.top,
                                bottom=thick_border.bottom
                            )
                            # Left/right borders for middle rows (already set by inner loops)

                        # Inner borders for all rows in section
                        for row_idx in range(start_row, end_row + 1):
                            for col_idx in range(1, 5):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                # Skip if this cell already has a border from outer box logic
                                if row_idx == start_row or row_idx == end_row:
                                    continue
                                # Apply thin borders on all sides
                                cell.border = thin_border

                    # Now apply the existing value‑cell coloring and recommendation highlighting
                    # (this must run after borders so it doesn't overwrite them)
                    for row in ws.iter_rows(min_row=2):
                        # Skip section header rows (already styled)
                        if str(row[0].value or '').startswith('──'):
                            continue
                        # Color the Value column (col 3, index 2)
                        value_cell = row[2]
                        val = value_cell.value
                        metric = str(row[1].value or '')
                        if isinstance(val, (int, float)):
                            if 'CRITICAL' in metric or val > 0 and ('Public' in metric or 'Without MFA' in metric or 'Open' in metric):
                                fill = PatternFill('solid', fgColor='FFB3B3')   # red
                            elif 'HIGH' in metric or ('Unencrypted' in metric or 'In ALARM' in metric):
                                fill = PatternFill('solid', fgColor='FFD9B3')   # orange
                            elif 'MEDIUM' in metric:
                                fill = PatternFill('solid', fgColor='FFFBB3')   # yellow
                            elif 'LOW' in metric or 'old' in metric.lower():
                                fill = PatternFill('solid', fgColor='D9F0B3')   # green
                            else:
                                fill = PatternFill('solid', fgColor='D6EAF8')   # info blue
                            value_cell.fill = fill
                        # Highlight Recommendation column (col 4, index 3) if text exists
                        rec_cell = row[3]
                        if rec_cell.value and str(rec_cell.value).strip():
                            rec_cell.fill = PatternFill('solid', fgColor='E6F0FA')
                            rec_cell.font = Font(bold=True)

                # ----- 3. EBS Snapshots: color only the Status cell -----
                if ws.title == 'EBS Snapshots':
                    # Find column index of 'Status'
                    headers = [cell.value for cell in ws[1]]
                    status_col_idx = None
                    for idx, h in enumerate(headers, start=1):
                        if h == 'Status':
                            status_col_idx = idx
                            break
                    if status_col_idx:
                        for row in ws.iter_rows(min_row=2):
                            status_cell = row[status_col_idx - 1]
                            if status_cell.value == 'Old (>90d)':
                                status_cell.fill = snapshot_old_fill
                            elif status_cell.value == 'Recent':
                                status_cell.fill = snapshot_recent_fill
               
                # ----- 3.1. IMDSv2 Only: color only the Status cell -----
                if ws.title == 'EC2 Instances':
                    # Find column index of 'IMDSv2 Only'
                    headers = [cell.value for cell in ws[1]]
                    imds_col_idx = None
                    for idx, h in enumerate(headers, start=1):
                        if h == 'IMDSv2 Only':
                            imds_col_idx = idx
                            break
                    if imds_col_idx:
                        fill_no = PatternFill('solid', fgColor='FFCCCC')  # light red
                        for row in ws.iter_rows(min_row=2):
                            cell = row[imds_col_idx - 1]
                            if cell.value == 'No':
                                cell.fill = fill_no

                # ----- 3.1. cloudfront: color only the Status cell -----
                if ws.title == 'CloudFront':
                    headers = [cell.value for cell in ws[1]]
                    enabled_col_idx = None
                    for idx, h in enumerate(headers, start=1):
                        if h == 'Enabled':
                            enabled_col_idx = idx
                            break
                    if enabled_col_idx:
                        fill_disabled = PatternFill('solid', fgColor='FFCCCC')
                        for row in ws.iter_rows(min_row=2):
                            cell = row[enabled_col_idx - 1]
                            if cell.value == False or cell.value == 'False':
                                cell.fill = fill_disabled
                # ----- 4. Cost Anomalies coloring (full row) -----
                if ws.title == 'Cost Anomalies':
                    for row in ws.iter_rows(min_row=2):
                        dev_val = None
                        for idx, cell in enumerate(row):
                            if ws.cell(row=1, column=idx+1).value == 'Deviation (USD)':
                                dev_val = cell.value
                                break
                        if dev_val is not None and isinstance(dev_val, (int, float)):
                            if dev_val > 0:
                                fill = PatternFill('solid', fgColor='FFCCCC')
                            elif dev_val < 0:
                                fill = PatternFill('solid', fgColor='CCFFCC')
                            else:
                                continue
                            for cell in row:
                                cell.fill = fill

                # ----- 5. Action‑required row coloring (whole row) for other sheets -----
                # Skip sheets already handled above
                if ws.title in ('Account Overview', 'Recommendations & Gaps', 'EBS Snapshots', 'Cost Anomalies'):
                    continue

                # Define rules for sheets where whole‑row highlighting is needed
                rules = {
                    'EBS Volumes': [('Attached To', 'Unattached')],
                    'Elastic IPs': [('Status', 'Unassociated')],
                    'Security Groups': [('Open to 0.0.0.0/0', 'Yes')],
                    'S3 Buckets': [('Public Access', 'Yes'), ('Public Access', 'Partial block')],
                    'RDS Instances': [
                        ('Publicly Accessible', 'Yes'),
                        ('Encrypted', 'No'),
                        ('Multi-AZ', 'No')
                    ],
                    'IAM Users': [
                        ('MFA Enabled', 'No'),
                        ('Status', 'Inactive'),
                        ('Old Keys (>90d)', lambda v: v > 0)
                    ],
                    'Load Balancers': [
                        ('Access Logs', 'false'),
                        ('Deletion Protection', 'false')
                    ],
                    'VPC Info': [('Flow Logs', 'Disabled')],
                    'EC2 Instances': [
                        ('IMDSv2 Only', 'No'),
                        ('Public IP', lambda v: v not in ('N/A', None, ''))
                    ],
                    'EBS Snapshots': [],   # already handled above
                    'AMIs': [('Age (days)', lambda v: isinstance(v, (int, float)) and v > 90)],
                    'AWS Backup Jobs': [('State', 'FAILED')],
                    'GuardDuty': [('High Severity Findings', lambda v: v > 0)],
                    'CloudWatch Alarms': [('State', 'ALARM')],
                }

                if ws.title in rules and rules[ws.title]:
                    headers = [cell.value for cell in ws[1]]
                    col_indices = {}
                    for idx, h in enumerate(headers, start=1):
                        if h:
                            col_indices[h] = idx

                    for row in ws.iter_rows(min_row=2):
                        should_highlight = False
                        for col_name, condition in rules[ws.title]:
                            if col_name not in col_indices:
                                continue
                            col_idx = col_indices[col_name]
                            cell_val = row[col_idx-1].value
                            if callable(condition):
                                if condition(cell_val):
                                    should_highlight = True
                                    break
                            else:
                                if str(cell_val) == condition:
                                    should_highlight = True
                                    break
                        if should_highlight:
                            for cell in row:
                                cell.fill = action_fill

        output.seek(0)
        return output.getvalue()

    def generate_service_list_bytes(self):
        print("Generating service list workbook...")
        services = []
        for sheet, rows in self.report_data.items():
            if not isinstance(rows, list) or not rows:
                continue
            service_name = sheet
            # Skip generic sheets not representing AWS services
            if service_name in ('Account Overview', 'Recommendations & Gaps', 'Cost Overview (raw)'):
                continue
            services.append({
                'Service Name': service_name,
                'Monitoring Recommendation': self._monitoring_recommendation_for_service(service_name),
                'Recommended Focus': self._service_focus_summary(service_name),
            })

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df = pd.DataFrame(services)
            if df.empty:
                df = pd.DataFrame([
                    {'Service Name': 'No detected service data', 'Monitoring Recommendation': '', 'Recommended Focus': ''}
                ])
            df.to_excel(writer, sheet_name='Service Monitor List', index=False)

            wb = writer.book
            ws = writer.sheets['Service Monitor List']
            for col in ws.columns:
                max_len = max((len(str(c.value or '')) for c in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            tab = Table(displayName='ServiceMonitorTable', ref=f'A1:C{len(df) + 1}')
            style = TableStyleInfo(
                name='TableStyleMedium9', showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False,
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)

        output.seek(0)
        return output.getvalue()

    def _monitoring_recommendation_for_service(self, service_name):
        recommendations = {
            'IAM Users': 'Monitor user logins, MFA changes, access key rotation, and inactive accounts.',
            'IAM Roles': 'Track role usage, permission changes, and role trust policy updates.',
            'EC2 Instances': 'Monitor instance health, public IP assignment, SSH exposure, and patch status.',
            'Security Groups': 'Watch open inbound rules, wide CIDR ranges, and changes to security group rules.',
            'EBS Volumes': 'Track unattached volumes, encryption state, and I/O anomalies.',
            'EBS Snapshots': 'Monitor old snapshots, retention policies, and public snapshot exposure.',
            'AMIs': 'Watch outdated AMIs and shared/public images.',
            'RDS Instances': 'Monitor public access, encryption, backups, and performance metrics.',
            'S3 Buckets': 'Track bucket public access, encryption, versioning, and logging settings.',
            'CloudWatch Alarms': 'Ensure alarms are active, firing, and routed to on-call channels.',
            'GuardDuty': 'Monitor high/critical findings and detection status active in all regions.',
            'CloudTrail Events': 'Track trail integrity, delivery failures, and suspicious events.',
            'KMS Keys': 'Watch key policy changes, key usage, and key rotation state.',
            'Load Balancers': 'Monitor listener configurations, HTTP/HTTPS exposure, and certificate expiration.',
            'Route53 Zones': 'Track hosted zone changes and public DNS record updates.',
            'CloudFront': 'Monitor distribution status, SSL cert expiry, and public content exposure.',
            'EKS Clusters': 'Watch control plane access and node group scaling/events.',
            'ECS Clusters': 'Monitor task health, container logs, and cluster resource usage.',
            'Secrets Manager': 'Track secret rotation state and unauthorized access attempts.',
        }
        return recommendations.get(service_name, 'Monitor changes and security posture for this service in your account.')

    def _service_focus_summary(self, service_name):
        focus = {
            'IAM Users': 'Focus on unused credentials, MFA enforcement, and access key expiry.',
            'IAM Roles': 'Focus on role privilege scope, external access, and session duration.',
            'EC2 Instances': 'Focus on publicly reachable instances and instance lifecycle management.',
            'Security Groups': 'Focus on rules allowing 0.0.0.0/0 and cross-account access.',
            'S3 Buckets': 'Focus on bucket ACLs, public policies, encryption, and logging.',
            'RDS Instances': 'Focus on public DB access, backups, and encryption settings.',
            'CloudWatch Alarms': 'Focus on alarms tied to critical resource health and security events.',
            'GuardDuty': 'Focus on high or suspicious findings and remediation timelines.',
            'CloudTrail Events': 'Focus on management event logging and log integrity failures.',
            'KMS Keys': 'Focus on key rotation, policy scope, and unused keys.',
            'Load Balancers': 'Focus on exposed listeners and certificate health.',
            'Route53 Zones': 'Focus on DNS changes and public record integrity.',
            'CloudFront': 'Focus on distribution public exposure and origin access configuration.',
        }
        return focus.get(service_name, 'Focus on key security and configuration risks for this service.')

    def generate_pdf_bytes(self, customer_name=None):
        from .pdf_report import generate_assessment_pdf
        return generate_assessment_pdf(self, customer_name=customer_name)
#--------------------------------------------------------------------
    def generate_html_report(self, output_path='aws_assessment_report.html'):
        """Generate a modern, animated HTML dashboard from collected data."""
        print("Generating HTML report...")

        # Helper to get account overview data (as a list of dicts)
        overview_data = self.report_data.get('Account Overview', [])
        findings = self.findings

        # Severity counts
        sev_counts = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0, 'INFO': 0}
        for f in findings:
            sev_counts[f['Severity']] += 1
        total_findings = len(findings)

        # Critical & high findings
        critical_high = [f for f in findings if f['Severity'] in ('CRITICAL', 'HIGH')]

        # Account info
        account_id = self.account_id
        account_name = self._account_name
        regions = len(self.regions)
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M UTC')

        # Build HTML
        html = f"""<!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>AWS Assessment Report – {account_name}</title>
        <style>
            * {{
                margin: 0;
                padding: 0;
                box-sizing: border-box;
            }}
            body {{
                font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
                background: #0a1628;
                color: #e2e8f0;
                padding: 2rem;
                animation: fadeIn 0.8s ease-out;
            }}
            @keyframes fadeIn {{
                from {{ opacity: 0; transform: translateY(20px); }}
                to {{ opacity: 1; transform: translateY(0); }}
            }}
            .container {{
                max-width: 1400px;
                margin: 0 auto;
            }}
            /* Header */
            .header {{
                background: linear-gradient(135deg, #0f2b3d, #1a4b6e);
                border-radius: 20px;
                padding: 2rem 2.5rem;
                margin-bottom: 2rem;
                box-shadow: 0 10px 30px rgba(0,0,0,0.3);
                display: flex;
                justify-content: space-between;
                align-items: center;
                flex-wrap: wrap;
                gap: 1rem;
            }}
            .header h1 {{
                font-size: 2rem;
                font-weight: 700;
            }}
            .header .sub {{
                color: #93c5fd;
                font-size: 1rem;
            }}
            .badge {{
                background: rgba(255,255,255,0.1);
                padding: 0.4rem 1rem;
                border-radius: 40px;
                font-size: 0.9rem;
                border: 1px solid rgba(255,255,255,0.2);
            }}
            /* Tabs */
            .tabs {{
                display: flex;
                gap: 0.5rem;
                margin-bottom: 1.5rem;
                flex-wrap: wrap;
            }}
            .tab-btn {{
                background: #142c44;
                border: 1px solid #1e3a5f;
                color: #93c5fd;
                padding: 0.6rem 1.5rem;
                border-radius: 40px;
                cursor: pointer;
                font-weight: 600;
                transition: all 0.2s;
            }}
            .tab-btn:hover {{
                background: #1e3a5f;
            }}
            .tab-btn.active {{
                background: #1d4ed8;
                border-color: #1d4ed8;
                color: white;
            }}
            .tab-content {{
                display: none;
                animation: fadeIn 0.5s ease;
            }}
            .tab-content.active {{
                display: block;
            }}
            /* Cards */
            .card {{
                background: #0f1f35;
                border: 1px solid #1e3a5f;
                border-radius: 16px;
                padding: 1.5rem;
                margin-bottom: 1.5rem;
                transition: transform 0.2s, box-shadow 0.2s;
            }}
            .card:hover {{
                transform: translateY(-4px);
                box-shadow: 0 12px 24px rgba(0,0,0,0.3);
            }}
            .card-title {{
                font-size: 1.2rem;
                font-weight: 700;
                margin-bottom: 1rem;
                color: #bfdbfe;
            }}
            /* Grid */
            .grid {{
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
                gap: 1rem;
            }}
            .stat {{
                background: #142c44;
                border-radius: 12px;
                padding: 1rem;
                text-align: center;
            }}
            .stat .number {{
                font-size: 2.2rem;
                font-weight: 700;
            }}
            .stat .label {{
                color: #7f9ec8;
                font-size: 0.8rem;
                text-transform: uppercase;
            }}
            .stat.critical .number {{ color: #3b82f6; }}
            .stat.high .number {{ color: #60a5fa; }}
            .stat.medium .number {{ color: #93c5fd; }}
            .stat.low .number {{ color: #bfdbfe; }}
            .stat.info .number {{ color: #dbeafe; }}
            /* Bar chart (severity) */
            .bar-container {{
                background: #1e3a5f;
                border-radius: 20px;
                overflow: hidden;
                height: 28px;
                margin: 0.3rem 0;
            }}
            .bar-fill {{
                height: 100%;
                border-radius: 20px;
                background: #3b82f6;
                transition: width 1s ease;
            }}
            .severity-item {{
                display: flex;
                align-items: center;
                gap: 0.5rem;
                margin-bottom: 0.5rem;
            }}
            .severity-item .label {{
                width: 80px;
                font-weight: 600;
            }}
            /* Findings table */
            .finding-table {{
                width: 100%;
                border-collapse: collapse;
                font-size: 0.9rem;
            }}
            .finding-table th {{
                background: #142c44;
                padding: 0.6rem;
                text-align: left;
                border-bottom: 2px solid #1e3a5f;
            }}
            .finding-table td {{
                padding: 0.6rem;
                border-bottom: 1px solid #1e3a5f;
            }}
            .finding-table tr:hover {{
                background: rgba(255,255,255,0.03);
            }}
            .severity-badge {{
                display: inline-block;
                padding: 0.2rem 0.6rem;
                border-radius: 20px;
                font-weight: 700;
                font-size: 0.7rem;
            }}
            .severity-badge.CRITICAL {{
                background: #3b82f6;
                color: white;
            }}
            .severity-badge.HIGH {{
                background: #60a5fa;
                color: #0a1628;
            }}
            .severity-badge.MEDIUM {{
                background: #93c5fd;
                color: #0a1628;
            }}
            .severity-badge.LOW {{
                background: #bfdbfe;
                color: #0a1628;
            }}
            .severity-badge.INFO {{
                background: #dbeafe;
                color: #0a1628;
            }}
            /* Account overview table */
            .overview-table {{
                width: 100%;
                border-collapse: collapse;
                font-size: 0.9rem;
            }}
            .overview-table th {{
                background: #142c44;
                padding: 0.5rem;
                text-align: left;
            }}
            .overview-table td {{
                padding: 0.4rem 0.5rem;
                border-bottom: 1px solid #1e3a5f;
            }}
            .overview-table .section-header {{
                background: #0f1f35;
                font-weight: 700;
                color: #bfdbfe;
            }}
            .footer {{
                text-align: center;
                margin-top: 2rem;
                color: #7f9ec8;
                font-size: 0.8rem;
                border-top: 1px solid #1e3a5f;
                padding-top: 1.5rem;
            }}
            @media (max-width: 768px) {{
                body {{ padding: 1rem; }}
                .header {{ flex-direction: column; align-items: flex-start; }}
            }}
        </style>
    </head>
    <body>
    <div class="container">
        <!-- Header -->
        <div class="header">
            <div>
                <h1>☁️ AWS Account Assessment</h1>
                <div class="sub">{account_name} · {account_id}</div>
                <div style="margin-top:0.5rem; font-size:0.9rem; color:#7f9ec8;">{regions} regions scanned · {now_str}</div>
            </div>
            <div>
                <span class="badge">🔴 {sev_counts['CRITICAL']} Critical</span>
                <span class="badge">🟠 {sev_counts['HIGH']} High</span>
                <span class="badge">🟡 {sev_counts['MEDIUM']} Medium</span>
                <span class="badge">🟢 {sev_counts['LOW']} Low</span>
                <span class="badge">🔵 {sev_counts['INFO']} Info</span>
            </div>
        </div>

        <!-- Tabs -->
        <div class="tabs">
            <button class="tab-btn active" data-tab="overview">📊 Overview</button>
            <button class="tab-btn" data-tab="vulns">🔍 Vulnerabilities</button>
            <button class="tab-btn" data-tab="details">📋 Account Details</button>
        </div>

        <!-- Tab: Overview -->
        <div id="overview" class="tab-content active">
            <div class="card">
                <div class="card-title">📈 Findings Summary</div>
                <div class="grid">
                    <div class="stat critical"><div class="number">{sev_counts['CRITICAL']}</div><div class="label">Critical</div></div>
                    <div class="stat high"><div class="number">{sev_counts['HIGH']}</div><div class="label">High</div></div>
                    <div class="stat medium"><div class="number">{sev_counts['MEDIUM']}</div><div class="label">Medium</div></div>
                    <div class="stat low"><div class="number">{sev_counts['LOW']}</div><div class="label">Low</div></div>
                    <div class="stat info"><div class="number">{sev_counts['INFO']}</div><div class="label">Info</div></div>
                    <div class="stat" style="background:#0f1f35;"><div class="number" style="color:#e2e8f0;">{total_findings}</div><div class="label">Total</div></div>
                </div>
                <!-- Severity bars -->
                <div style="margin-top:1.5rem;">
                    <div class="severity-item"><span class="label">Critical</span><div class="bar-container"><div class="bar-fill" style="width:{ (sev_counts['CRITICAL']/total_findings*100) if total_findings else 0 }%; background:#3b82f6;"></div></div></div>
                    <div class="severity-item"><span class="label">High</span><div class="bar-container"><div class="bar-fill" style="width:{ (sev_counts['HIGH']/total_findings*100) if total_findings else 0 }%; background:#60a5fa;"></div></div></div>
                    <div class="severity-item"><span class="label">Medium</span><div class="bar-container"><div class="bar-fill" style="width:{ (sev_counts['MEDIUM']/total_findings*100) if total_findings else 0 }%; background:#93c5fd;"></div></div></div>
                    <div class="severity-item"><span class="label">Low</span><div class="bar-container"><div class="bar-fill" style="width:{ (sev_counts['LOW']/total_findings*100) if total_findings else 0 }%; background:#bfdbfe;"></div></div></div>
                    <div class="severity-item"><span class="label">Info</span><div class="bar-container"><div class="bar-fill" style="width:{ (sev_counts['INFO']/total_findings*100) if total_findings else 0 }%; background:#dbeafe;"></div></div></div>
                </div>
            </div>
        </div>

        <!-- Tab: Vulnerabilities (Critical & High only) -->
        <div id="vulns" class="tab-content">
            <div class="card">
                <div class="card-title">🚨 Critical & High Findings ({len(critical_high)})</div>
                {f'''
                <table class="finding-table">
                    <thead><tr><th>Severity</th><th>Category</th><th>Resource</th><th>Issue</th><th>Recommendation</th></tr></thead>
                    <tbody>
                        {''.join(f'''
                        <tr>
                            <td><span class="severity-badge {f['Severity']}">{f['Severity']}</span></td>
                            <td>{f['Category']}</td>
                            <td>{f['Resource']}</td>
                            <td>{f['Issue']}</td>
                            <td>{f['Recommendation']}</td>
                        </tr>
                        ''' for f in critical_high[:50])}
                    </tbody>
                </table>
                ''' if critical_high else '<p style="color:#7f9ec8;">✅ No critical or high findings.</p>'}
                {f'<p style="margin-top:0.5rem; color:#7f9ec8;">Showing up to 50 findings. Full list in Excel report.</p>' if len(critical_high)>50 else ''}
            </div>
        </div>

        <!-- Tab: Account Details -->
        <div id="details" class="tab-content">
            <div class="card">
                <div class="card-title">📋 Account Overview</div>
                {f'''
                <table class="overview-table">
                    <thead><tr><th>Category</th><th>Metric</th><th>Value</th><th>Recommendation</th></tr></thead>
                    <tbody>
                        {''.join(f'''
                        <tr class="{'section-header' if str(row.get('Category','')).startswith('──') else ''}">
                            <td>{row.get('Category','')}</td>
                            <td>{row.get('Metric','')}</td>
                            <td>{row.get('Value','')}</td>
                            <td>{row.get('Recommendation','')}</td>
                        </tr>
                        ''' for row in overview_data[:100])}
                    </tbody>
                </table>
                ''' if overview_data else '<p>No overview data available.</p>'}
            </div>
        </div>

        <div class="footer">
            Generated by AWS Assessment Script · Data collected from {regions} regions · Full details in Excel report.
        </div>
    </div>

    <script>
        // Tab switching with animation
        document.querySelectorAll('.tab-btn').forEach(btn => {{
            btn.addEventListener('click', function() {{
                const tabId = this.getAttribute('data-tab');
                document.querySelectorAll('.tab-content').forEach(tc => tc.classList.remove('active'));
                document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
                document.getElementById(tabId).classList.add('active');
                this.classList.add('active');
            }});
        }});
    </script>
    </body>
    </html>
    """
        # Write the file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html)
        print(f"✅ HTML report saved to {output_path}")
    # ─────────────────────────────────────────────────────────────────────────
    # RUNNER
    # ─────────────────────────────────────────────────────────────────────────
    def run_assessment(self, progress_callback=None):
        print(f"\n{'='*60}")
        print(f"  AWS FULL ACCOUNT ASSESSMENT  |  Account: {self.account_id}")
        print(f"  Regions: {len(self.regions)}  |  Max workers: {MAX_WORKERS}")
        print(f"{'='*60}\n")

        collectors = [
            self.get_iam_users_data,
            self.get_iam_password_policy,
            self.get_iam_roles,
            self.get_ec2_instances,
            self.get_security_groups,
            self.get_ebs_volumes,
            self.get_ebs_snapshots,
            self.get_amis,                          # your AMI collector
            self.get_rds_instances,
            self.get_aurora_clusters,               # new
            self.get_dynamodb_tables,               # new
            self.get_elasticache_clusters,          # new
            self.get_cloudfront,                    # new
            self.get_s3_buckets,
            self.get_cloudtrail_events,
            self.get_cloudwatch_alarms,
            self.get_guardduty_status,
            self.get_waf_status,                    # already there
            self.get_vpc_info,
            self.get_load_balancers,
            self.get_elastic_ips,
            self.get_route53_zones,
            self.get_vpn_connections,
            self.get_aws_backup_status,
            self.get_cost_overview,
            self.get_secrets_manager,
            self.get_opensearch_domains,
            self.get_redshift_clusters,
            self.get_eks_clusters,
            self.get_ecs_clusters,
            self.get_kms_keys,
        ]

        # Top-level: all collectors run in parallel
        import time
        t0 = time.time()
        total = len(collectors)
        completed = 0
        if progress_callback:
            progress_callback('Initializing collectors', 0, total)
        with ThreadPoolExecutor(max_workers=total) as executor:
            futures = {executor.submit(fn): fn.__name__ for fn in collectors}
            for future in as_completed(futures):
                name = futures[future]
                completed += 1
                try:
                    records = future.result()
                    print(f"  ✔ {name} → {records} records")
                except Exception as e:
                    records = 0
                    print(f"  ✘ {name} failed: {e}")
                if progress_callback:
                    progress_callback(name, completed, total)
        print(f"\n  ⏱  Data collection: {time.time() - t0:.1f}s")

        self.build_account_overview()

        sev = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0, 'INFO': 0}
        for f in self.findings:
            sev[f['Severity']] = sev.get(f['Severity'], 0) + 1

        print(f"\n{'='*60}")
        print(f"  COMPLETE  |  Findings: {len(self.findings)}")
        print(f"  🔴 {sev['CRITICAL']}  🟠 {sev['HIGH']}  🟡 {sev['MEDIUM']}  🟢 {sev['LOW']}  🔵 {sev['INFO']}")
        print(f"{'='*60}\n")


# ── Lambda handler ────────────────────────────────────────────────────────────
def lambda_handler(event, context):
    s3_bucket = os.environ.get('S3_BUCKET')
    s3_prefix = os.environ.get('S3_PREFIX', 'assessments')
    if not s3_bucket:
        return {'statusCode': 400, 'body': json.dumps('S3_BUCKET env var not set')}
    try:
        a = AWSAssessment()
        a.run_assessment()
        excel_bytes = a.generate_excel_bytes()
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
        safe_account_name = a.account_name.replace(' ', '_').replace('/', '_')
        s3_key = f"{s3_prefix}/aws_assessment_{safe_account_name}_{timestamp}.xlsx"
        boto3.client('s3').put_object(
            Bucket=s3_bucket, Key=s3_key, Body=excel_bytes,
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        return {'statusCode': 200, 'body': json.dumps({'s3_location': f"s3://{s3_bucket}/{s3_key}"})}
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'statusCode': 500, 'body': json.dumps({'error': str(e)})}


# ── Local runner ──────────────────────────────────────────────────────────────
if __name__ == '__main__':
    a = AWSAssessment()
    a.run_assessment()
    print("run_assessment() completed")
    a.generate_html_report('aws_assessment_report.html')   # <-- new line
    excel_bytes = a.generate_excel_bytes()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    safe_account_name = a.account_name.replace(' ', '_').replace('/', '_')
    output_path = os.path.join(script_dir, f"aws_assessment_{safe_account_name}_{timestamp}.xlsx")
    with open(output_path, 'wb') as f:
        f.write(excel_bytes)
    print(f"✅ Report saved → {output_path}")